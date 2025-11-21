# routes.py
from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from flask_login import login_user, logout_user, login_required, current_user
from extensions import db, mail
from flask_mail import Message
from werkzeug.utils import secure_filename
from models import Valor, Nominacion, Usuario, CicloEscolar
from utils import generar_pdf, ciclo_actual
import pandas as pd
import tempfile
import os
from io import BytesIO
from models import Maestro
from models import Alumno, Bloque
from flask import send_file
from datetime import datetime
from flask import send_file, request, jsonify
from flask import jsonify
from datetime import datetime
from utils import admin_required
from utils import cerrar_eventos_vencidos

# -------------------------------
# 🔹 Definición de Blueprints
# -------------------------------
nom = Blueprint('nom', __name__)
admin_bp = Blueprint('admin_bp', __name__, url_prefix='/admin')


# -------------------------------
# 🔹 Ruta Keep-Alive (para evitar que Render se duerma)
# -------------------------------
@nom.route("/keepalive")
def keepalive():
    return "OK", 200

@nom.route("/status")
def status():
    return "OK", 200


# -------------------------------
# 🔹 Página principal (protegida)
# -------------------------------
@nom.route('/')
@login_required
def principal():
    if current_user.rol == 'admin':
        return render_template('panel_admin.html')
    elif current_user.rol == 'profesor':
        return redirect(url_for('nom.panel_profesor'))
    else:
        return render_template('principal.html')  # si en el futuro hay alumnos



# -------------------------------
# 🔹 Login y logout
# -------------------------------
@nom.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password = request.form['password']
        user = Usuario.query.filter_by(email=email).first()

        if user and user.check_password(password):
            login_user(user)
            flash(f'Bienvenido, {user.nombre}', 'success')
            return redirect(url_for('nom.principal'))
        else:
            flash('Correo o contraseña incorrectos', 'danger')

    return render_template('login.html')


@nom.route('/logout')
@login_required
def logout():
    logout_user()
    flash("Sesión cerrada correctamente.", "success")
    return redirect(url_for('nom.login'))


# -------------------------------
# 🔹 Panel de administración general
# -------------------------------
@nom.route('/admin')
@login_required
def panel_admin():
    if current_user.rol != 'admin':
        flash("🚫 Se necesita ser administrador para poder acceder a este apartado.", "danger")
        return redirect(url_for('nom.principal'))
    return render_template('panel_admin.html')



# ======================================================
# 🧠 Función auxiliar: actualiza a EXCELENCIA automáticamente
# ======================================================
def actualizar_a_excelencia(alumno_id, ciclo_id):
    """Convierte las nominaciones de un alumno a EXCELENCIA si ya alcanzó 3."""
    from models import Nominacion, Valor
    from sqlalchemy.orm import joinedload
    import datetime

    # Buscar valor EXCELENCIA o crearlo si no existe (ignorando mayúsculas)
    excelencia = Valor.query.filter(
        db.func.lower(Valor.nombre) == "excelencia",
        Valor.ciclo_id == ciclo_id
    ).first()
    if not excelencia:
        excelencia = Valor(nombre="EXCELENCIA", ciclo_id=ciclo_id, activo=True)
        db.session.add(excelencia)
        db.session.commit()

    # Si ya tiene EXCELENCIA, no hacer nada
    tiene_excelencia = (
        Nominacion.query.filter_by(alumno_id=alumno_id, ciclo_id=ciclo_id)
        .join(Valor)
        .filter(db.func.lower(Valor.nombre) == "excelencia")
        .first()
    )
    if tiene_excelencia:
        return False  # ya tenía excelencia

    # Traer todas las nominaciones actuales (sin contar EXCELENCIA)
    nominaciones = (
        Nominacion.query
        .options(joinedload(Nominacion.valor))
        .filter(
            Nominacion.alumno_id == alumno_id,
            Nominacion.ciclo_id == ciclo_id
        )
        .all()
    )

    # Filtrar solo las nominaciones normales
    nominaciones_normales = [n for n in nominaciones if n.valor and n.valor.nombre.upper() != "EXCELENCIA"]

    if len(nominaciones_normales) < 3:
        return False  # aún no alcanza 3

    # Concatenar comentarios y valores previos
    comentarios = []
    valores = []
    for n in nominaciones_normales:
        if n.valor:
            valores.append(n.valor.nombre)
        if n.comentario:
            comentarios.append(n.comentario.strip())

    texto_final = f"Valores obtenidos: {', '.join(valores)}. " \
                    f"Comentarios: {' | '.join(comentarios)}"

    # 🔹 Marcar las tres nominaciones previas como "solo visuales" (no exportables)
    for n in nominaciones_normales:
        if n.valor and n.valor.nombre.upper() != "EXCELENCIA":
            if "[EXCELENCIA-VISUAL]" not in (n.comentario or ""):
                n.comentario = (n.comentario or "") + " [EXCELENCIA-VISUAL]"
    db.session.commit()

    # 🔹 Crear nueva nominación de EXCELENCIA
    nueva = Nominacion(
        alumno_id=alumno_id,
        maestro_id=nominaciones_normales[-1].maestro_id if nominaciones_normales else None,
        valor_id=excelencia.id,
        ciclo_id=ciclo_id,
        comentario=texto_final,
        evento_id=nominaciones_normales[-1].evento_id if nominaciones_normales else None,
        tipo="alumno",
        fecha=datetime.datetime.utcnow()
    )
    db.session.add(nueva)
    db.session.commit()

    return True

# ======================================================
# 🧠 Verificar reversión de EXCELENCIA si bajan de 3 nominaciones
# ======================================================
def verificar_reversion_excelencia(alumno_id, ciclo_id):
    """Elimina la nominación de EXCELENCIA si el alumno baja de 3 nominaciones normales."""
    from models import Nominacion, Valor

    excelencia = Valor.query.filter(
        db.func.lower(Valor.nombre) == "excelencia",
        Valor.ciclo_id == ciclo_id
    ).first()
    if not excelencia:
        return

    # Contar nominaciones normales (sin excelencia)
    nominaciones_normales = (
        Nominacion.query
        .join(Valor)
        .filter(
            Nominacion.alumno_id == alumno_id,
            Nominacion.ciclo_id == ciclo_id,
            db.func.lower(Valor.nombre) != "excelencia"
        ).count()
    )

    # Si bajó de 3 → eliminar la EXCELENCIA
    if nominaciones_normales < 3:
        nom_excelencia = (
            Nominacion.query
            .filter_by(alumno_id=alumno_id, ciclo_id=ciclo_id, valor_id=excelencia.id)
            .first()
        )
        if nom_excelencia:
            db.session.delete(nom_excelencia)
            db.session.commit()

# ======================================================
# 🔁 Recalcular comentario de la nominación EXCELENCIA
# ======================================================
def recalcular_comentario_excelencia(alumno_id, ciclo_id):
    """
    Reconstruye el comentario de la nominación EXCELENCIA a partir de las
    3 nominaciones 'visuales' (las que incluyen el tag [EXCELENCIA-VISUAL]).
    Si no existe EXCELENCIA, no hace nada.
    """
    from models import Nominacion, Valor

    excelencia = Valor.query.filter(
        db.func.lower(Valor.nombre) == "excelencia",
        Valor.ciclo_id == ciclo_id
    ).first()
    if not excelencia:
        return

    # Buscar la nominación de EXCELENCIA (si existe)
    nom_excelencia = (
        Nominacion.query
        .filter_by(alumno_id=alumno_id, ciclo_id=ciclo_id, valor_id=excelencia.id)
        .first()
    )
    if not nom_excelencia:
        return  # aún no hay excelencia para este alumno

    # Tomar las nominaciones 'visuales' (las 3 que originaron la excelencia)
    visuales = (
        Nominacion.query.join(Valor)
        .filter(
            Nominacion.alumno_id == alumno_id,
            Nominacion.ciclo_id == ciclo_id,
            db.func.lower(Valor.nombre) != "excelencia",
            Nominacion.comentario.ilike("%[EXCELENCIA-VISUAL]%")
        )
        .order_by(Nominacion.fecha.asc())
        .all()
    )

    # Limpiar y reconstruir valores y comentarios
    valores = []
    comentarios = []
    for n in visuales:
        # re-apegar el tag por si lo quitaron al editar
        if "[EXCELENCIA-VISUAL]" not in (n.comentario or ""):
            n.comentario = (n.comentario or "").strip() + " [EXCELENCIA-VISUAL]"
        if n.valor:
            valores.append(n.valor.nombre)
        # comentario sin el tag para el texto final
        if n.comentario:
            comentarios.append(n.comentario.replace("[EXCELENCIA-VISUAL]", "").strip())

    # Armar comentario final (soporta si hay <3 o >3 por algún motivo)
    texto_final = ""
    if valores:
        texto_final += f"Valores obtenidos: {', '.join(valores)}. "
    if comentarios:
        texto_final += f"Comentarios: {' | '.join(comentarios)}"

    nom_excelencia.comentario = texto_final.strip()
    db.session.commit()



from werkzeug.security import generate_password_hash  # asegúrate de tenerlo arriba

@nom.route('/crear_usuario', methods=['GET', 'POST'])
@login_required
@admin_required
def crear_usuario():
    if request.method == 'POST':
        nombre = request.form['nombre']
        correo = request.form['correo']
        password = request.form['password']
        rol = request.form.get('rol', 'admin')  # si no se envía, por defecto admin

        # Verificar si ya existe
        existente = Usuario.query.filter_by(email=correo).first()
        if existente:
            flash("⚠️ Ya existe un usuario con ese correo.", "warning")
            return redirect(url_for('nom.crear_usuario'))

        # ✅ Crear usuario correctamente
        nuevo = Usuario(nombre=nombre, email=correo, rol=rol)
        nuevo.set_password(password)
        db.session.add(nuevo)
        db.session.commit()

        flash(f"✅ Usuario '{nombre}' creado correctamente como {rol}.", "success")
        # 🔹 En lugar de redirigir a otra página, recargamos la misma
        return redirect(url_for('nom.crear_usuario'))

    # Para GET
    return render_template('crear_usuario.html', rol_predefinido='admin')

# -------------------------------
# 🔹 Importar usuarios desde Excel o CSV
# -------------------------------
UPLOAD_FOLDER = tempfile.gettempdir()
ALLOWED_EXTENSIONS = {'xlsx', 'csv'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


@nom.route('/admin/usuarios/importar/<rol>', methods=['POST'])
@login_required
def importar_usuarios(rol):
    if current_user.rol != 'admin':
        flash("🚫 No tienes permisos para importar usuarios.", "danger")
        return redirect(url_for('nom.panel_usuarios'))

    file = request.files.get('file')
    if not file or file.filename == '':
        flash("⚠️ No se seleccionó ningún archivo.", "warning")
        return redirect(url_for('nom.panel_usuarios'))

    if not allowed_file(file.filename):
        flash("❌ Solo se permiten archivos Excel (.xlsx) o CSV.", "danger")
        return redirect(url_for('nom.panel_usuarios'))

    filename = secure_filename(file.filename)
    filepath = os.path.join(UPLOAD_FOLDER, filename)
    file.save(filepath)

    try:
        if filename.endswith('.xlsx'):
            df = pd.read_excel(filepath)
        else:
            df = pd.read_csv(filepath)

        columnas_requeridas = {'nombre', 'email', 'contraseña'}
        if not columnas_requeridas.issubset(set(df.columns.str.lower())):
            flash("❌ La plantilla no contiene las columnas requeridas (nombre, email, contraseña).", "danger")
            return redirect(url_for('nom.panel_usuarios'))

        usuarios_creados = 0
        for _, row in df.iterrows():
            nombre = str(row['nombre']).strip()
            email = str(row['email']).strip().lower()
            password = str(row['contraseña']).strip()

            if not nombre or not email or not password:
                continue

            existente = Usuario.query.filter_by(email=email).first()
            if existente:
                continue

            nuevo_usuario = Usuario(nombre=nombre, email=email, rol=rol)
            nuevo_usuario.set_password(password)
            db.session.add(nuevo_usuario)
            usuarios_creados += 1

            if usuarios_creados % 50 == 0:
                db.session.commit()

        db.session.commit()
        flash(f"✅ Se importaron correctamente {usuarios_creados} usuarios del rol {rol}.", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error al procesar el archivo: {str(e)}", "danger")

    finally:
        if os.path.exists(filepath):
            os.remove(filepath)

    return redirect(url_for('nom.panel_usuarios'))


# -------------------------------
# 🔹 Panel de usuarios
# -------------------------------
@nom.route('/admin/usuarios')
@login_required
def panel_usuarios():
    if current_user.rol != 'admin':
        flash("🚫 Solo los administradores pueden acceder a esta sección.", "danger")
        return redirect(url_for('nom.principal'))
    return render_template('panel_usuarios.html')


@nom.route('/admin/usuarios/plantilla/<rol>')
@login_required
def descargar_plantilla(rol):
    if current_user.rol != 'admin':
        flash("🚫 Solo los administradores pueden descargar plantillas.", "danger")
        return redirect(url_for('nom.panel_usuarios'))

    columnas = ["nombre", "email", "contraseña"]
    df = pd.DataFrame(columns=columnas)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name=f"{rol.capitalize()}s")
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"Plantilla_{rol}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@nom.route('/admin/usuarios/lista/<rol>')
@login_required
def lista_usuarios(rol):
    if current_user.rol != 'admin':
        return jsonify({"error": "No autorizado"}), 403

    usuarios = Usuario.query.filter_by(rol=rol).all()
    data = [
        {"id": u.id, "nombre": u.nombre, "email": u.email, "rol": u.rol}
        for u in usuarios
    ]
    return jsonify(data)



# -------------------------------
# 🔹 Ruta de prueba de conexión a Neon
# -------------------------------
@nom.route("/check_db")
def check_db():
    try:
        db.session.execute("SELECT 1")
        return "✅ Conectado correctamente a Neon.tech"
    except Exception as e:
        return f"❌ Error de conexión: {e}"


# -------------------------------
# 🔹 Administración de ciclos escolares
# -------------------------------
@admin_bp.route('/ciclos', methods=['GET', 'POST'])
def gestionar_ciclos():
    ciclos = CicloEscolar.query.order_by(CicloEscolar.id.desc()).all()
    if request.method == 'POST':
        nombre = request.form['nombre']
        inicio = request.form.get('fecha_inicio')
        fin = request.form.get('fecha_fin')
        nuevo = CicloEscolar(nombre=nombre, fecha_inicio=inicio, fecha_fin=fin)
        db.session.add(nuevo)
        db.session.commit()
        flash(f'Ciclo {nombre} creado correctamente.')
        return redirect(url_for('admin_bp.gestionar_ciclos'))
    return render_template('admin_ciclos.html', ciclos=ciclos)


# -------------------------------
# 🔹 Activar un ciclo escolar
# -------------------------------
@admin_bp.route('/ciclos/activar/<int:ciclo_id>', methods=['POST'])
def activar_ciclo(ciclo_id):
    ciclo = CicloEscolar.query.get_or_404(ciclo_id)

    # Desactivar todos los demás
    CicloEscolar.query.update({CicloEscolar.activo: False})

    # Activar este
    ciclo.activo = True
    db.session.commit()

    return jsonify({"success": True, "activo_id": ciclo_id})

# -------------------------------
# 🔹 Importar maestros al ciclo activo
# -------------------------------
# -------------------------------
# 🔹 Importar maestros al ciclo activo (CSV o Excel)
# -------------------------------
@admin_bp.route('/maestros/importar', methods=['POST'])
@login_required
def importar_maestros_ciclo():
    if current_user.rol != 'admin':
        flash("🚫 Solo los administradores pueden importar maestros.", "danger")
        return redirect(url_for('nom.panel_usuarios'))

    file = request.files.get('file')
    if not file or file.filename == '':
        flash("⚠️ No se seleccionó ningún archivo.", "warning")
        return redirect(url_for('admin_bp.maestros_ciclo'))

    filename = secure_filename(file.filename)
    filepath = os.path.join(tempfile.gettempdir(), filename)
    file.save(filepath)

    try:
        # ✅ Detección automática del tipo de archivo
        ext = filename.split('.')[-1].lower()

        if ext in ['xlsx', 'xls']:
            # Forzamos que lea contraseñas como texto
            df = pd.read_excel(filepath, dtype=str)
        elif ext in ['csv']:
            df = pd.read_csv(filepath, dtype=str, encoding='utf-8')
        else:
            flash("❌ Formato no soportado. Usa un archivo .xlsx o .csv", "danger")
            return redirect(url_for('admin_bp.maestros_ciclo'))

        # ✅ Normalizar columnas
        df.columns = df.columns.str.strip().str.lower()

        # Aceptar ambas versiones de 'contraseña' o 'contrasena'
        if 'contrasena' in df.columns and 'contraseña' not in df.columns:
            df.rename(columns={'contrasena': 'contraseña'}, inplace=True)

        columnas_requeridas = {'nombre', 'email', 'contraseña'}
        if not columnas_requeridas.issubset(df.columns):
            flash("❌ La plantilla debe contener las columnas: nombre, email y contraseña.", "danger")
            return redirect(url_for('admin_bp.maestros_ciclo'))

        # ✅ Obtener ciclo activo
        ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()
        if not ciclo_activo:
            flash("⚠️ No hay ningún ciclo activo. Activa un ciclo primero.", "warning")
            return redirect(url_for('admin_bp.gestionar_ciclos'))

        importados, existentes = 0, 0

        # ✅ Iterar y limpiar datos
        for _, row in df.iterrows():
            nombre = str(row.get('nombre', '')).strip()
            correo = str(row.get('email', '')).strip().lower()
            raw_pass = row.get('contraseña', '')

            # 🔍 Convertir contraseñas numéricas a texto limpio
            if pd.isna(raw_pass):
                password = ''
            else:
                password = str(raw_pass).strip()
                # Si viene en formato "123456.0" → "123456"
                if password.endswith('.0'):
                    password = password[:-2]
                # Si viene como número grande sin comillas, asegúrate que sea string
                if password.isdigit():
                    password = password

            if not nombre or not correo:
                continue

            # Crear o actualizar usuario
            usuario = Usuario.query.filter_by(email=correo).first()
            if not usuario:
                usuario = Usuario(nombre=nombre, email=correo, rol='profesor')
                usuario.set_password(password if password else '123456')
                db.session.add(usuario)
            else:
                if password:
                    usuario.set_password(password)

            # Crear maestro solo si no existe ya
            existe_maestro = Maestro.query.filter_by(correo=correo, ciclo_id=ciclo_activo.id).first()
            if existe_maestro:
                existentes += 1
                continue

            maestro = Maestro(nombre=nombre, correo=correo, ciclo_id=ciclo_activo.id)
            db.session.add(maestro)
            importados += 1

        db.session.commit()
        flash(f"✅ {importados} maestros importados correctamente al ciclo {ciclo_activo.nombre}. {existentes} ya existían.", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error al importar: {e}", "danger")

    finally:
        if os.path.exists(filepath):
            os.remove(filepath)

    return redirect(url_for('admin_bp.maestros_ciclo'))


# -------------------------------
# 🔹 Listado de maestros del ciclo activo
# -------------------------------
@admin_bp.route('/maestros')
@login_required
def maestros_ciclo():
    if current_user.rol != 'admin':
        flash("🚫 Solo los administradores pueden acceder.", "danger")
        return redirect(url_for('nom.principal'))

    ciclo = CicloEscolar.query.filter_by(activo=True).first()
    maestros = Maestro.query.filter_by(ciclo_id=ciclo.id).all() if ciclo else []
    return render_template('admin_maestros.html', maestros=maestros, ciclo=ciclo)

# -------------------------------
# 🔹 Listado e importación de alumnos del ciclo activo
# -------------------------------
@admin_bp.route('/alumnos', methods=['GET'])
@login_required
def alumnos_ciclo():
    if current_user.rol != 'admin':
        flash("🚫 Solo los administradores pueden acceder.", "danger")
        return redirect(url_for('nom.principal'))

    ciclo = CicloEscolar.query.filter_by(activo=True).first()
    alumnos = Alumno.query.filter_by(ciclo_id=ciclo.id).all() if ciclo else []

    # 🔹 Ordena los bloques numéricamente si se llaman "Bloque 1", "Bloque 2", etc.
    bloques = []
    if ciclo:
        bloques = (
            Bloque.query
            .filter_by(ciclo_id=ciclo.id)
            .order_by(
                db.cast(db.func.substr(Bloque.nombre, 8), db.Integer).asc()
            )
            .all()
        )

    return render_template('admin_alumnos.html', alumnos=alumnos, ciclo=ciclo, bloques=bloques)

# -------------------------------
# 🔹 Importar alumnos al ciclo activo (normalizado y sin duplicados)
# -------------------------------
@admin_bp.route('/alumnos/importar', methods=['POST'])
@login_required
def importar_alumnos_ciclo():
    if current_user.rol != 'admin':
        flash("🚫 Solo los administradores pueden importar alumnos.", "danger")
        return redirect(url_for('nom.principal'))

    file = request.files.get('file')
    if not file or file.filename == '':
        flash("⚠️ No se seleccionó ningún archivo.", "warning")
        return redirect(url_for('admin_bp.alumnos_ciclo'))

    filename = secure_filename(file.filename)
    filepath = os.path.join(tempfile.gettempdir(), filename)
    file.save(filepath)

    try:
        # Detectar formato (Excel o CSV)
        df = pd.read_excel(filepath) if filename.endswith('.xlsx') else pd.read_csv(filepath)
        df.columns = [c.strip().lower() for c in df.columns]  # Normaliza encabezados

        columnas_requeridas = {'nombre', 'grado', 'grupo', 'nivel', 'bloque'}
        if not columnas_requeridas.issubset(set(df.columns)):
            faltan = ", ".join(columnas_requeridas - set(df.columns))
            flash(f"❌ La plantilla debe contener las columnas: {faltan}", "danger")
            return redirect(url_for('admin_bp.alumnos_ciclo'))

        ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()
        if not ciclo_activo:
            flash("⚠️ No hay un ciclo activo. Activa un ciclo primero.", "warning")
            return redirect(url_for('admin_bp.gestionar_ciclos'))

        importados, existentes = 0, 0

        for _, row in df.iterrows():
            nombre = str(row["nombre"]).strip().title()
            grado = str(row["grado"]).strip()
            grupo = str(row["grupo"]).strip().upper()
            nivel = str(row["nivel"]).strip().capitalize()
            bloque_nombre = str(row["bloque"]).strip()

            if not (nombre and grado and grupo and nivel and bloque_nombre):
                continue

            # Buscar o crear bloque
            bloque = Bloque.query.filter_by(nombre=bloque_nombre, ciclo_id=ciclo_activo.id).first()
            if not bloque:
                bloque = Bloque(nombre=bloque_nombre, ciclo_id=ciclo_activo.id)
                db.session.add(bloque)
                db.session.commit()

            # Evita duplicados ignorando mayúsculas/minúsculas
            ya = Alumno.query.filter(
                db.func.lower(Alumno.nombre) == nombre.lower(),
                Alumno.grado == grado,
                db.func.upper(Alumno.grupo) == grupo.upper(),
                Alumno.ciclo_id == ciclo_activo.id
            ).first()

            if ya:
                existentes += 1
                continue

            nuevo = Alumno(
                nombre=nombre,
                grado=grado,
                grupo=grupo,
                nivel=nivel,
                ciclo_id=ciclo_activo.id,
                bloque_id=bloque.id
            )
            db.session.add(nuevo)
            importados += 1

        db.session.commit()
        flash(f"✅ {importados} alumnos importados al ciclo {ciclo_activo.nombre}. {existentes} ya existían.", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error al importar alumnos: {e}", "danger")

    finally:
        if os.path.exists(filepath):
            os.remove(filepath)

    return redirect(url_for('admin_bp.alumnos_ciclo'))


# -------------------------------
# 🔹 Gestión de Bloques del Ciclo Activo
# -------------------------------
# ------------------------------------------------------------
# 🧩 Gestión completa de Bloques Académicos del Ciclo Activo
# ------------------------------------------------------------
@admin_bp.route('/bloques', methods=['GET', 'POST'])
@login_required
def bloques_ciclo():
    # ✅ Validar acceso solo para administradores
    if current_user.rol != 'admin':
        flash("🚫 Solo los administradores pueden acceder.", "danger")
        return redirect(url_for('nom.principal'))

    # ✅ Obtener ciclo activo
    ciclo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo:
        flash("⚠️ No hay un ciclo activo. Activa un ciclo antes de administrar bloques.", "warning")
        return redirect(url_for('admin_bp.gestionar_ciclos'))

    # ✅ Crear nuevo bloque
    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        if not nombre:
            flash("⚠️ El nombre del bloque no puede estar vacío.", "warning")
            return redirect(url_for('admin_bp.bloques_ciclo'))

        existente = Bloque.query.filter_by(nombre=nombre, ciclo_id=ciclo.id).first()
        if existente:
            flash(f"🚫 El bloque '{nombre}' ya existe en este ciclo.", "danger")
        else:
            nuevo = Bloque(nombre=nombre, ciclo_id=ciclo.id)
            db.session.add(nuevo)
            db.session.commit()
            flash(f"✅ Bloque '{nombre}' creado exitosamente.", "success")

        return redirect(url_for('admin_bp.bloques_ciclo'))

    # ✅ Obtener lista de bloques del ciclo
    bloques = (
        db.session.query(Bloque)
        .filter_by(ciclo_id=ciclo.id)
        .order_by(Bloque.orden.asc())
        .all()
    )

    # ✅ Contar alumnos por bloque
    conteos = {b.id: Alumno.query.filter_by(bloque_id=b.id).count() for b in bloques}

    # ✅ Armar vista con grados y grupos por bloque
    data = []
    for bloque in bloques:
        alumnos = Alumno.query.filter_by(bloque_id=bloque.id).all()
        if alumnos:
            grados_dict = {}
            for a in alumnos:
                grados_dict.setdefault(a.grado, set()).add(a.grupo)
            grados = [{'grado': g, 'grupos': sorted(list(grs))} for g, grs in grados_dict.items()]
        else:
            grados = []
        data.append({'bloque': bloque, 'grados': grados})

    # ✅ Renderizar plantilla moderna unificada
    return render_template(
        'admin_bloques.html',
        ciclo=ciclo,
        bloques=bloques,
        conteos=conteos,
        data=data
    )


# -------------------------------
# 🔹 Alumnos del ciclo activo (vista + importación por BLOQUE)
# -------------------------------
ALLOWED_EXT_ALUMNOS = {"xlsx", "csv"}

def _allowed_file_alumnos(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXT_ALUMNOS


@admin_bp.route('/administradores', methods=['GET', 'POST'])
@login_required
@admin_required
def admin_administradores():
    return render_template('admin_administradores.html')

# -------------------------------
# 🟢 Importar alumnos al bloque seleccionado
# -------------------------------
@admin_bp.route('/alumnos/importar_bloque', methods=['POST'])
@login_required
def importar_alumnos_por_bloque():
    """Importa alumnos al ciclo activo según bloque seleccionado"""
    if current_user.rol != 'admin':
        flash("🚫 Solo los administradores pueden importar alumnos.", "danger")
        return redirect(url_for('nom.principal'))

    # ✅ Verificar ciclo activo
    ciclo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo:
        flash("⚠️ No hay un ciclo activo. Activa un ciclo primero.", "warning")
        return redirect(url_for('admin_bp.gestionar_ciclos'))

    # ✅ Verificar bloque seleccionado
    bloque_id = request.form.get("bloque_id", type=int)
    if not bloque_id:
        flash("⚠️ Selecciona un bloque para importar.", "warning")
        return redirect(url_for('admin_bp.alumnos_ciclo'))

    bloque = Bloque.query.filter_by(id=bloque_id, ciclo_id=ciclo.id).first()
    if not bloque:
        flash("❌ El bloque seleccionado no existe en el ciclo activo.", "danger")
        return redirect(url_for('admin_bp.alumnos_ciclo'))

    # ✅ Verificar archivo
    file = request.files.get("file")
    if not file or file.filename == "":
        flash("⚠️ No se seleccionó ningún archivo.", "warning")
        return redirect(url_for('admin_bp.alumnos_ciclo'))

    if not _allowed_file_alumnos(file.filename):
        flash("❌ Solo se permiten archivos .xlsx o .csv", "danger")
        return redirect(url_for('admin_bp.alumnos_ciclo'))

    # ✅ Guardar temporalmente el archivo
    filename = secure_filename(file.filename)
    tmp_path = os.path.join(tempfile.gettempdir(), filename)
    file.save(tmp_path)

    try:
        # Leer archivo
        df = pd.read_excel(tmp_path) if filename.lower().endswith(".xlsx") else pd.read_csv(tmp_path)
        df.columns = [c.strip().lower() for c in df.columns]

        # ✅ Validar columnas requeridas
        requeridas = {"nombre", "grado", "grupo", "nivel"}
        if not requeridas.issubset(set(df.columns)):
            faltan = ", ".join(sorted(requeridas - set(df.columns)))
            flash(f"❌ Faltan columnas requeridas: {faltan}", "danger")
            return redirect(url_for('admin_bp.alumnos_ciclo'))

        importados, existentes = 0, 0

        # ✅ Iterar e importar alumnos
        for _, row in df.iterrows():
            nombre = str(row.get("nombre", "")).strip()
            grado = str(row.get("grado", "")).strip()
            grupo = str(row.get("grupo", "")).strip()
            nivel = str(row.get("nivel", "")).strip()

            if not (nombre and grado and grupo and nivel):
                continue

            existente = Alumno.query.filter_by(
                nombre=nombre, grado=grado, grupo=grupo, ciclo_id=ciclo.id
            ).first()

            if existente:
                existentes += 1
                continue

            nuevo = Alumno(
                nombre=nombre,
                grado=grado,
                grupo=grupo,
                nivel=nivel,
                ciclo_id=ciclo.id,
                bloque_id=bloque.id
            )
            db.session.add(nuevo)
            importados += 1

        db.session.commit()
        flash(f"✅ {importados} alumnos importados a «{bloque.nombre}». {existentes} ya existían.", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error al importar: {str(e)}", "danger")

    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)

    return redirect(url_for('admin_bp.alumnos_ciclo'))


# -------------------------------
# 🔹 Descargar plantilla Excel (alumnos por bloque)
# -------------------------------
@admin_bp.route('/alumnos/plantilla')
@login_required
def plantilla_alumnos_bloque():
    """Genera y descarga una plantilla Excel con columnas básicas"""
    try:
        # Crear DataFrame vacío con columnas correctas
        import pandas as pd
        from io import BytesIO

        columnas = ["nombre", "grado", "grupo", "nivel"]
        df = pd.DataFrame(columns=columnas)

        # Escribir en memoria (sin guardar en disco)
        output = BytesIO()
        with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
            df.to_excel(writer, index=False, sheet_name="AlumnosPorBloque")
        output.seek(0)

        # Enviar al navegador como archivo descargable
        return send_file(
            output,
            as_attachment=True,
            download_name="Plantilla_alumnos_por_bloque.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        flash(f"❌ Error al generar la plantilla: {e}", "danger")
        return redirect(url_for('admin_bp.admin_alumnos'))


# -------------------------------
# 🔹 Editar alumno
# -------------------------------
@admin_bp.route('/alumnos/editar/<int:id>', methods=['POST'])
@login_required
def editar_alumno(id):
    if current_user.rol != 'admin':
        return jsonify({"error": "No autorizado"}), 403

    alumno = Alumno.query.get_or_404(id)
    data = request.get_json()

    alumno.nombre = data.get("nombre", alumno.nombre).title()
    alumno.grado = data.get("grado", alumno.grado)
    alumno.grupo = data.get("grupo", alumno.grupo).upper()
    alumno.nivel = data.get("nivel", alumno.nivel).capitalize()

    db.session.commit()
    return jsonify({"message": f"Alumno {alumno.nombre} actualizado con éxito."}), 200


# -------------------------------
# 🔹 Eliminar alumno
# -------------------------------
@admin_bp.route('/alumnos/eliminar/<int:id>', methods=['DELETE'])
@login_required
def eliminar_alumno(id):
    if current_user.rol != 'admin':
        return jsonify({"error": "No autorizado"}), 403

    alumno = Alumno.query.get_or_404(id)
    db.session.delete(alumno)
    db.session.commit()
    return jsonify({"message": f"Alumno {alumno.nombre} eliminado."}), 200


# -------------------------------
# 🔹 Decorador para rutas solo de administrador
# -------------------------------
from functools import wraps
from flask import abort

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or current_user.rol != 'admin':
            flash("🚫 Solo los administradores pueden acceder a esta sección.", "danger")
            return redirect(url_for('nom.principal'))
        return f(*args, **kwargs)
    return decorated_function


# -------------------------------
# 🔹 VALORES POR CICLO ESCOLAR
# -------------------------------
@admin_bp.route('/valores', methods=['GET'])
@login_required
def listar_valores():
    # Usar tu modelo correcto
    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo_activo:
        flash("⚠️ No hay ciclo activo.", "warning")
        return redirect(url_for('admin_bp.gestionar_ciclos'))

    valores = Valor.query.filter_by(ciclo_id=ciclo_activo.id).all()
    # 👇 Aquí corregimos la ruta del template
    return render_template('valores.html', valores=valores, ciclo=ciclo_activo)



@admin_bp.route('/valores/nuevo', methods=['POST'])
@login_required
def crear_valor():
    nombre = request.form.get('nombre', '').strip().title()
    descripcion = request.form.get('descripcion', '').strip()
    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()

    if not ciclo_activo:
        flash("⚠️ No hay ciclo activo.", "danger")
        return redirect(url_for('admin_bp.listar_valores'))

    if Valor.query.filter_by(nombre=nombre, ciclo_id=ciclo_activo.id).first():
        flash("⚠️ Ya existe un valor con ese nombre en este ciclo.", "warning")
        return redirect(url_for('admin_bp.listar_valores'))

    nuevo = Valor(nombre=nombre, descripcion=descripcion, ciclo_id=ciclo_activo.id)
    db.session.add(nuevo)
    db.session.commit()
    flash("✅ Valor agregado exitosamente.", "success")
    return redirect(url_for('admin_bp.listar_valores'))


@admin_bp.route('/valores/<int:id>/toggle', methods=['POST'])
@login_required
@admin_required
def alternar_valor(id):
    """Activa o desactiva un valor existente."""
    valor = Valor.query.get_or_404(id)
    valor.activo = not valor.activo
    db.session.commit()
    flash(f"🔁 Valor {'activado' if valor.activo else 'desactivado'}.", "info")
    return redirect(url_for('admin_bp.listar_valores'))


@admin_bp.route('/valores/<int:id>/eliminar', methods=['POST'])
@login_required
@admin_required
def eliminar_valor(id):
    """Elimina un valor del ciclo actual."""
    valor = Valor.query.get_or_404(id)
    db.session.delete(valor)
    db.session.commit()
    flash("🗑️ Valor eliminado correctamente.", "success")
    return redirect(url_for('admin_bp.listar_valores'))

@admin_bp.route('/nominaciones', methods=['GET'])
@login_required
@admin_required
def gestionar_nominaciones():
    """Vista de administración para revisar todas las nominaciones del ciclo activo"""
    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo_activo:
        flash("⚠️ No hay ciclo escolar activo.", "warning")
        return redirect(url_for('admin_bp.gestionar_ciclos'))

    # Traer todas las nominaciones ordenadas por fecha
    nominaciones = (
        Nominacion.query
        .filter_by(ciclo_id=ciclo_activo.id)
        .order_by(Nominacion.fecha.desc())
        .all()
    )

    # Dividir por tipo
    nominaciones_alumnos = [n for n in nominaciones if n.tipo == "alumno"]
    nominaciones_personal = [n for n in nominaciones if n.tipo == "personal"]

    return render_template(
        'admin_nominaciones.html',
        ciclo=ciclo_activo,
        nominaciones_alumnos=nominaciones_alumnos,
        nominaciones_personal=nominaciones_personal
    )



# Maestro nomina a alumnos (sin dependencia de bloque)
@nom.route('/nominaciones/alumno', methods=['GET', 'POST'])
@login_required
def nominar_alumno():
    cerrar_eventos_vencidos()
    # 1️⃣ Validar rol del usuario
    if current_user.rol != 'profesor':
        flash("🚫 Solo los profesores pueden registrar nominaciones.", "danger")
        return redirect(url_for('nom.principal'))

    # 2️⃣ Validar que haya ciclo activo
    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo_activo:
        flash("⚠️ No hay ciclo activo disponible.", "warning")
        return redirect(url_for('nom.principal'))

    # 3️⃣ Obtener maestro actual
    maestro = Maestro.query.filter_by(correo=current_user.email, ciclo_id=ciclo_activo.id).first()
    if not maestro:
        flash("⚠️ No se encontró tu registro como maestro en el ciclo activo.", "warning")
        return redirect(url_for('nom.principal'))

    # 4️⃣ Obtener evento activo del BLOQUE del maestro
    evento_abierto = (
        EventoAsamblea.query
        .filter_by(ciclo_id=ciclo_activo.id, activo=True)
        .order_by(EventoAsamblea.fecha_evento.asc())
        .first()
    )

    # 🔍 Validar que exista un evento activo y que esté en tiempo
    if not evento_abierto:
        flash("⚠️ No hay un evento activo configurado para tu bloque.", "warning")
        return redirect(url_for('nom.principal'))

    # 👉 Si el evento está marcado como activo pero su fecha de cierre ya pasó, se debe bloquear también
    if not evento_abierto.esta_abierto:
        # Desactivamos el evento automáticamente (opcional, pero útil)
        evento_abierto.activo = False
        db.session.commit()

        flash(
            f"🚫 El evento {evento_abierto.nombre_mes} ya cerró las nominaciones "
            f"(cierre: {evento_abierto.fecha_cierre_nominaciones.strftime('%d/%m/%Y %H:%M')}).",
            "danger"
        )
        return redirect(url_for('nom.principal'))


    # 5️⃣ Procesar envío del formulario
    if request.method == 'POST':
        valor_id = request.form.get('valor_id')
        alumno_ids = request.form.getlist('alumnos')
        comentario = request.form.get('comentario', '').strip()

        if not valor_id or not alumno_ids:
            flash("⚠️ Selecciona al menos un alumno y un valor.", "warning")
            return redirect(url_for('nom.nominar_alumno'))

        nuevas = 0
        for alumno_id in alumno_ids:
            # Evitar duplicados: mismo maestro, alumno y valor en el mismo ciclo
            existente = Nominacion.query.filter_by(
                alumno_id=alumno_id,
                maestro_id=maestro.id,
                valor_id=valor_id,
                ciclo_id=ciclo_activo.id
            ).first()
            if existente:
                continue

            nueva_nom = Nominacion(
                alumno_id=alumno_id,
                maestro_id=maestro.id,
                valor_id=valor_id,
                ciclo_id=ciclo_activo.id,
                comentario=comentario,
                evento_id=evento_abierto.id,
                tipo='alumno'
            )
            db.session.add(nueva_nom)
            nuevas += 1

        db.session.commit()
        flash(f"✅ Se registraron {nuevas} nominaciones al evento {evento_abierto.nombre_mes}.", "success")
        return redirect(url_for('nom.nominar_alumno'))

    # 6️⃣ Cargar datos para el formulario
    alumnos = (
        Alumno.query
        .filter_by(ciclo_id=ciclo_activo.id)
        .order_by(Alumno.grado, Alumno.grupo, Alumno.nombre)
        .all()
    )
    valores = Valor.query.filter_by(ciclo_id=ciclo_activo.id, activo=True).all()

    # 7️⃣ Mostrar también las nominaciones previas del maestro
    nominaciones = (
        Nominacion.query
        .filter_by(ciclo_id=ciclo_activo.id, maestro_id=maestro.id)
        .order_by(Nominacion.fecha.desc())
        .all()
    )

    return render_template(
        'nominacion_alumno.html',
        alumnos=alumnos,
        valores=valores,
        nominaciones=nominaciones,
        ciclo=ciclo_activo,
        evento=evento_abierto
    )


# 🔧 Función para asegurar que los admins existan como maestros en el ciclo actual
def sincronizar_admins_como_maestros(ciclo_activo):
    from models import Usuario, Maestro, db
    admins = Usuario.query.filter_by(rol='admin').all()
    for admin in admins:
        existente = Maestro.query.filter_by(correo=admin.email, ciclo_id=ciclo_activo.id).first()
        if not existente:
            nuevo = Maestro(
                nombre=admin.nombre,
                correo=admin.email,
                ciclo_id=ciclo_activo.id,
                activo=True
            )
            db.session.add(nuevo)
    db.session.commit()
# Maestro nomina a otro maestro (sin bloque, solo requiere evento activo)
# ============================================================
# 🔧 FUNCIÓN AUXILIAR: sincronizar administradores como maestros
# ============================================================
def sincronizar_admins_como_maestros(ciclo_activo):
    """Asegura que todos los administradores tengan un registro en Maestro."""
    from models import Usuario, Maestro, db

    admins = Usuario.query.filter_by(rol='admin').all()
    for admin in admins:
        # Verificar si ya existe en la tabla maestros para este ciclo
        existente = Maestro.query.filter_by(correo=admin.email, ciclo_id=ciclo_activo.id).first()
        if not existente:
            nuevo = Maestro(
                nombre=admin.nombre,
                correo=admin.email,
                ciclo_id=ciclo_activo.id,
                activo=True
            )
            db.session.add(nuevo)
    db.session.commit()


# ============================================================
# 👨‍🏫 NOMINAR PERSONAL (profesor o admin)
# ============================================================
@nom.route('/nominaciones/personal', methods=['GET', 'POST'])
@login_required
def nominar_personal():
    cerrar_eventos_vencidos()
    """Vista para que los profesores (o admins) nominen a otros miembros del personal."""
    # 1️⃣ Validar rol del usuario
    if current_user.rol not in ['profesor', 'admin']:
        flash("🚫 Solo los profesores o administradores pueden registrar nominaciones.", "danger")
        return redirect(url_for('nom.principal'))

    # 2️⃣ Validar ciclo activo
    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo_activo:
        flash("⚠️ No hay ciclo activo disponible.", "warning")
        return redirect(url_for('nom.principal'))

    # 🔹 Asegurar que los administradores también estén en la tabla maestros
    sincronizar_admins_como_maestros(ciclo_activo)

    # 3️⃣ Obtener maestro actual
    maestro = Maestro.query.filter_by(correo=current_user.email, ciclo_id=ciclo_activo.id, activo=True).first()
    if not maestro:
        flash("⚠️ No se encontró tu registro como maestro activo en el ciclo actual.", "warning")
        return redirect(url_for('nom.principal'))

    # 4️⃣ Obtener evento activo del ciclo (sin bloque)
    evento_abierto = (
        EventoAsamblea.query
        .filter_by(ciclo_id=ciclo_activo.id, activo=True)
        .order_by(EventoAsamblea.fecha_evento.asc())
        .first()
    )

    if not evento_abierto or not evento_abierto.esta_abierto:
        flash("🚫 No hay un evento de asamblea abierto para nominaciones en este momento.", "warning")
        return redirect(url_for('nom.principal'))

    # 5️⃣ Procesar nominaciones (POST)
    if request.method == 'POST':
        valor_id = request.form.get('valor_id')
        nominados = request.form.getlist('maestros')
        comentario = request.form.get('comentario', '').strip()

        if not valor_id or not nominados:
            flash("⚠️ Selecciona al menos un maestro y un valor.", "warning")
            return redirect(url_for('nom.nominar_personal'))

        nuevas = 0
        duplicados = []

        for nominado_id in nominados:
            existente = Nominacion.query.filter_by(
                maestro_nominado_id=nominado_id,
                maestro_id=maestro.id,
                valor_id=valor_id,
                ciclo_id=ciclo_activo.id
            ).first()

            if existente:
                nominado = Maestro.query.get(nominado_id)
                if nominado:
                    duplicados.append(nominado.nombre)
                continue

            nueva_nom = Nominacion(
                maestro_nominado_id=nominado_id,
                maestro_id=maestro.id,
                valor_id=valor_id,
                ciclo_id=ciclo_activo.id,
                comentario=comentario,
                evento_id=evento_abierto.id,
                tipo='personal'
            )
            db.session.add(nueva_nom)
            nuevas += 1

        db.session.commit()

        # ✅ Mensajes más detallados
        if nuevas > 0 and not duplicados:
            flash(f"✅ Se registraron {nuevas} nominaciones de personal al evento {evento_abierto.nombre_mes}.", "success")

        elif nuevas > 0 and duplicados:
            lista = ', '.join(duplicados)
            flash(f"✅ Se registraron {nuevas} nominaciones nuevas. ⚠️ Ya habías nominado a: {lista}.", "warning")

        elif nuevas == 0 and duplicados:
            lista = ', '.join(duplicados)
            flash(f"⚠️ No se registraron nuevas nominaciones porque ya nominaste a {lista} con este valor.", "warning")

        else:
            flash("⚠️ No se registraron nominaciones.", "warning")

        return redirect(url_for('nom.nominar_personal'))

    # 6️⃣ Datos para el formulario (GET)
    maestros = Maestro.query.filter(
        Maestro.ciclo_id == ciclo_activo.id,
        Maestro.id != maestro.id,
        Maestro.activo == True
    ).order_by(Maestro.nombre.asc()).all()

    valores = Valor.query.filter_by(ciclo_id=ciclo_activo.id, activo=True).all()

    nominaciones = (
        Nominacion.query
        .filter_by(ciclo_id=ciclo_activo.id, maestro_id=maestro.id)
        .order_by(Nominacion.fecha.desc())
        .all()
    )

    valores_json = [{"id": v.id, "nombre": v.nombre} for v in valores]

    return render_template(
        'nominacion_personal.html',
        maestros=maestros,
        valores=valores,
        valores_json=valores_json,
        nominaciones=nominaciones,
        ciclo=ciclo_activo,
        evento=evento_abierto
    )


@nom.route('/admin/nominaciones/export', methods=['GET'])
@login_required
def exportar_nominaciones_excel():
    """
    Exporta nominaciones del ciclo activo a Excel.
    ?tipo=alumno|personal  (opcional; si no se envía, exporta todo)
    """
    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo_activo:
        return jsonify({"error": "No hay ciclo escolar activo."}), 400

    q = Nominacion.query.filter_by(ciclo_id=ciclo_activo.id)
    tipo = request.args.get('tipo')
    if tipo in ('alumno', 'personal'):
        q = q.filter(Nominacion.tipo == tipo)

    q = q.order_by(Nominacion.fecha.desc()).all()

    filas = []
    for n in q:
        filas.append({
            "Fecha": n.fecha.strftime("%Y-%m-%d %H:%M") if n.fecha else "",
            "Tipo": n.tipo or "",
            "Valor": n.valor.nombre if getattr(n, 'valor', None) else "",
            "Maestro (autor)": n.maestro.nombre if getattr(n, 'maestro', None) else "",
            "Alumno nominado": n.alumno.nombre if getattr(n, 'alumno', None) else "",
            "Personal nominado": n.maestro_nominado.nombre if getattr(n, 'maestro_nominado', None) else "",
            "Comentario": (n.comentario or n.razon) if hasattr(n, 'comentario') or hasattr(n, 'razon') else "",
        })

    df = pd.DataFrame(filas)
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Nominaciones')
        ws = writer.sheets['Nominaciones']
        for i, col in enumerate(df.columns):
            width = 12
            if not df.empty:
                width = min(60, max(12, df[col].astype(str).map(len).max() + 2))
            ws.set_column(i, i, width)
    output.seek(0)

    nombre = f"nominaciones_{tipo or 'todas'}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output,
        as_attachment=True,
        download_name=nombre,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
@nom.route('/admin/nominaciones/data', methods=['GET'])
@login_required
def obtener_nominaciones_data():
    """Devuelve nominaciones del ciclo activo filtradas por parámetros opcionales."""
    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo_activo:
        return jsonify({"items": [], "error": "No hay ciclo activo."})

    tipo = request.args.get('tipo')  # alumno | personal
    valor_id = request.args.get('valor_id', type=int)
    maestro_id = request.args.get('maestro_id', type=int)
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')

    q = Nominacion.query.filter(Nominacion.ciclo_id == ciclo_activo.id)

    if tipo in ('alumno', 'personal'):
        q = q.filter(Nominacion.tipo == tipo)
    if valor_id:
        q = q.filter(Nominacion.valor_id == valor_id)
    if maestro_id:
        q = q.filter(Nominacion.maestro_id == maestro_id)

    if fecha_desde:
        try:
            fd = datetime.strptime(fecha_desde, "%Y-%m-%d")
            q = q.filter(Nominacion.fecha >= fd)
        except ValueError:
            pass

    if fecha_hasta:
        try:
            fh = datetime.strptime(fecha_hasta, "%Y-%m-%d")
            fh = fh.replace(hour=23, minute=59, second=59)
            q = q.filter(Nominacion.fecha <= fh)
        except ValueError:
            pass

    nominaciones = q.order_by(Nominacion.fecha.desc()).all()

    datos = []
    for n in nominaciones:
        datos.append({
            "id": n.id,
            "fecha": n.fecha.strftime("%d/%m/%Y"),
            "tipo": n.tipo,
            "valor": n.valor.nombre if n.valor else "",
            "maestro": n.maestro.nombre if n.maestro else "",
            "alumno": n.alumno.nombre if getattr(n, 'alumno', None) else "",
            "maestro_nominado": n.maestro_nominado.nombre if getattr(n, 'maestro_nominado', None) else "",
            "comentario": n.comentario or "",
        })

    return jsonify({"items": datos, "total": len(datos)})

@nom.route('/admin/catalogo/valores')
@login_required
def catalogo_valores():
    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()
    valores = Valor.query.filter_by(ciclo_id=ciclo_activo.id).order_by(Valor.nombre.asc()).all()
    return jsonify([{"id": v.id, "nombre": v.nombre} for v in valores])

@nom.route('/admin/catalogo/maestros')
@login_required
def catalogo_maestros():
    maestros = Maestro.query.order_by(Maestro.nombre.asc()).all()
    return jsonify([{"id": m.id, "nombre": m.nombre} for m in maestros])


@nom.route('/admin/usuarios/lista/alumnos')
@login_required
def lista_alumnos():
    """Devuelve JSON con alumnos del ciclo activo."""
    if current_user.rol != 'admin':
        return jsonify({"error": "No autorizado"}), 403

    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo_activo:
        return jsonify({"error": "No hay ciclo activo."}), 400

    alumnos = Alumno.query.filter_by(ciclo_id=ciclo_activo.id).order_by(
        Alumno.grado, Alumno.grupo, Alumno.nombre
    ).all()

    data = [
        {
            "id": a.id,
            "nombre": a.nombre,
            "grado": a.grado,
            "grupo": a.grupo,
            "nivel": a.nivel
        }
        for a in alumnos
    ]
    return jsonify(data)


# -------------------------------
# 🔹 Editar alumno (vista y actualización)
# -------------------------------
@admin_bp.route('/alumnos/editar/<int:id>', methods=['GET'])
@login_required
def editar_alumno_vista(id):
    if current_user.rol != 'admin':
        flash("🚫 Solo los administradores pueden acceder.", "danger")
        return redirect(url_for('nom.principal'))

    alumno = Alumno.query.get_or_404(id)
    return render_template('editar_alumno.html', alumno=alumno)


@admin_bp.route('/alumnos/editar/<int:id>', methods=['POST'])
@login_required
def actualizar_alumno(id):
    if current_user.rol != 'admin':
        flash("🚫 Solo los administradores pueden editar alumnos.", "danger")
        return redirect(url_for('nom.principal'))

    alumno = Alumno.query.get_or_404(id)
    alumno.nombre = request.form.get('nombre', alumno.nombre).strip().title()
    alumno.grado = request.form.get('grado', alumno.grado).strip()
    alumno.grupo = request.form.get('grupo', alumno.grupo).strip().upper()
    alumno.nivel = request.form.get('nivel', alumno.nivel).strip().capitalize()

    db.session.commit()
    flash(f"✅ Alumno {alumno.nombre} actualizado correctamente.", "success")
    return redirect(url_for('nom.panel_usuarios'))

# -------------------------------
# 🔹 Editar usuario (admin o profesor)
# -------------------------------
@nom.route('/admin/usuarios/editar/<int:id>', methods=['GET'])
@login_required
def editar_usuario_vista(id):
    if current_user.rol != 'admin':
        flash("🚫 Solo los administradores pueden acceder a esta sección.", "danger")
        return redirect(url_for('admin_bp.maestros_ciclo'))

    usuario = Usuario.query.get_or_404(id)
    return render_template('editar_usuario.html', usuario=usuario)


@nom.route('/admin/usuarios/editar/<int:id>', methods=['POST'])
@login_required
def actualizar_usuario(id):
    if current_user.rol != 'admin':
        flash("🚫 No autorizado.", "danger")
        return redirect(url_for('nom.principal'))

    usuario = Usuario.query.get_or_404(id)
    usuario.nombre = request.form.get('nombre', usuario.nombre).strip().title()
    usuario.email = request.form.get('email', usuario.email).strip().lower()
    usuario.rol = request.form.get('rol', usuario.rol)

    db.session.commit()
    flash(f"✅ Usuario {usuario.nombre} actualizado correctamente.", "success")
    return redirect(url_for('nom.panel_usuarios'))


# -------------------------------
# 🔹 Editar profesor (vista + actualización)
# -------------------------------
# ============================================================
# ✏️ Editar PROFESOR (GET muestra formulario / POST guarda)
# ============================================================

@admin_bp.route('/maestros/lista')
@login_required
def lista_maestros_json():
    # Solo admin
    if current_user.rol != 'admin':
        return jsonify({"error": "No autorizado"}), 403

    ciclo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo:
        return jsonify([])

    maes = Maestro.query.filter_by(ciclo_id=ciclo.id).order_by(Maestro.nombre.asc()).all()
    data = [{"id": m.id, "nombre": m.nombre, "email": m.correo} for m in maes]
    return jsonify(data)

@admin_bp.route('/maestros/editar/<int:id>', methods=['GET'])
@login_required
def editar_maestro_vista(id):
    if current_user.rol != 'admin':
        flash("🚫 Solo los administradores.", "danger")
        return redirect(url_for('nom.panel_usuarios'))

    maestro = Maestro.query.get_or_404(id)
    return render_template('editar_profesor.html', maestro=maestro)

@admin_bp.route('/maestros/editar/<int:id>', methods=['POST'])
@login_required
def actualizar_maestro(id):
    if current_user.rol != 'admin':
        flash("🚫 Solo los administradores.", "danger")
        return redirect(url_for('nom.panel_usuarios'))

    maestro = Maestro.query.get_or_404(id)
    maestro.nombre = request.form.get('nombre', maestro.nombre).strip()
    maestro.correo = request.form.get('correo', maestro.correo).strip().lower()
    db.session.commit()
    flash("✅ Maestro actualizado correctamente.", "success")
    return redirect(url_for('admin_bp.maestros_ciclo'))

@admin_bp.route('/maestros/eliminar/<int:id>', methods=['POST'])
@login_required
def eliminar_maestro(id):
    """Desactiva o reactiva un maestro sin eliminarlo."""
    if current_user.rol != 'admin':
        flash("🚫 Solo los administradores pueden realizar esta acción.", "danger")
        return redirect(url_for('admin_bp.maestros_ciclo'))

    maestro = Maestro.query.get_or_404(id)

    # Alternar estado
    maestro.activo = not maestro.activo
    db.session.commit()

    estado = "activado" if maestro.activo else "desactivado"
    flash(f"✅ El maestro '{maestro.nombre}' ha sido {estado}.", "success")

    return redirect(url_for('admin_bp.maestros_ciclo'))

# --- DASHBOARD JSON ---
# -------------------------------
# 🔹 Dashboard del Administrador

from datetime import datetime

@admin_bp.route('/dashboard')
@login_required
def admin_dashboard():
    cerrar_eventos_vencidos()

    if current_user.rol != 'admin':
        flash("🚫 Solo los administradores pueden acceder al dashboard.", "danger")
        return redirect(url_for('nom.principal'))

    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo_activo:
        flash("⚠️ No hay ciclo escolar activo.", "warning")
        return redirect(url_for('nom.panel_admin'))

    # 🔎 Traer todos los eventos del ciclo
    eventos_ciclo = (
        EventoAsamblea.query
        .filter_by(ciclo_id=ciclo_activo.id)
        .order_by(EventoAsamblea.mes_ordinal, EventoAsamblea.fecha_evento)
        .all()
    )
    
        # 🧩 NUEVO: traer los bloques del ciclo activo
    bloques = (
        Bloque.query
        .filter_by(ciclo_id=ciclo_activo.id)
        .order_by(db.cast(db.func.substr(Bloque.nombre, 8), db.Integer))
        .all()
    )

    # 📦 Agrupar por nombre de mes
    meses = {}
    for e in eventos_ciclo:
        meses.setdefault(e.nombre_mes, []).append(e)

    # 🧠 Construir "eventos_unicos": un representante por mes + flag esta_abierto_mes
    now = datetime.utcnow()
    eventos_unicos = []
    for nombre_mes, lista in meses.items():
        # ✅ usar fecha_cierre_nominaciones
        algun_abierto = any(
            (ev.activo is True) and (
                ev.fecha_cierre_nominaciones is None or ev.fecha_cierre_nominaciones >= now
            )
            for ev in lista
        )

        rep = lista[0]                       # representativo del mes (ya ordenado)
        rep.esta_abierto_mes = algun_abierto # flag para el template
        eventos_unicos.append(rep)

    # 🟢 Mes por defecto: el más reciente con algún bloque abierto; si no, el primero
    abiertos = [ev for ev in eventos_unicos if getattr(ev, 'esta_abierto_mes', False)]
    if abiertos:
        seleccionado = max(abiertos, key=lambda ev: ev.fecha_evento or now)
        mes_seleccionado = seleccionado.nombre_mes
    else:
        mes_seleccionado = eventos_unicos[0].nombre_mes if eventos_unicos else None

    return render_template(
        'admin_dashboard.html',
        eventos=eventos_unicos,
        bloques=bloques,
        mes_seleccionado=mes_seleccionado
    )

# 🔹 Datos JSON para el dashboard (filtrados por mes)
@admin_bp.route('/dashboard/data/resumen')
@login_required
def data_dashboard_resumen():
    if current_user.rol != 'admin':
        return jsonify({"error": "Solo los administradores pueden consultar este recurso."}), 403

    ciclo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo:
        return jsonify({"error": "No hay ciclo activo."}), 400

    mes = request.args.get("mes")  # p.ej. ?mes=Octubre

    # Base query: nominaciones del ciclo
    query = Nominacion.query.filter_by(ciclo_id=ciclo.id)
    if mes:
        query = query.join(EventoAsamblea).filter(EventoAsamblea.nombre_mes == mes)

    nominaciones = query.all()

    # KPIs
    total_maestros = Maestro.query.filter_by(ciclo_id=ciclo.id).count()
    total_alumnos = Alumno.query.filter_by(ciclo_id=ciclo.id).count()
    total_nominaciones = len(nominaciones)

    # Nominaciones por tipo
    por_tipo = {"alumno": 0, "personal": 0}
    for n in nominaciones:
        if n.tipo in por_tipo:
            por_tipo[n.tipo] += 1

    # Top valores
    conteo_valores = {}
    for n in nominaciones:
        if n.valor:
            nombre = n.valor.nombre
            conteo_valores[nombre] = conteo_valores.get(nombre, 0) + 1
    por_valor = [{"valor": v, "n": c} for v, c in sorted(conteo_valores.items(), key=lambda x: x[1], reverse=True)]

    # Nominaciones por día (solo días del mes si se pasó ?mes=)
    fechas = {}
    for n in nominaciones:
        fecha_str = n.fecha.strftime("%d/%m/%Y") if n.fecha else "Sin fecha"
        fechas[fecha_str] = fechas.get(fecha_str, 0) + 1
    por_dia = [{"fecha": f, "n": c} for f, c in sorted(fechas.items())]

    return jsonify({
        "totales": {
            "maestros": total_maestros,
            "alumnos": total_alumnos,
            "nominaciones": total_nominaciones
        },
        "por_tipo": por_tipo,
        "por_valor": por_valor,
        "por_dia": por_dia
    })


# -------------------------------
# 🔹 Endpoints JSON para gráficas
# -------------------------------
@admin_bp.route('/dashboard/data/resumen')
@login_required
def dashboard_data_resumen():
    if current_user.rol != 'admin':
        return jsonify({"error": "Solo administradores"}), 403

    ciclo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo:
        return jsonify({"error": "No hay ciclo activo"}), 400

    total_maestros = Maestro.query.filter_by(ciclo_id=ciclo.id).count()
    total_alumnos = Alumno.query.filter_by(ciclo_id=ciclo.id).count()
    total_nom = Nominacion.query.filter_by(ciclo_id=ciclo.id).count()

    por_tipo = (
        db.session.query(Nominacion.tipo, db.func.count(Nominacion.id))
        .filter(Nominacion.ciclo_id == ciclo.id)
        .group_by(Nominacion.tipo)
        .all()
    )
    por_tipo_dict = {t: c for t, c in por_tipo}

    por_valor = (
        db.session.query(Valor.nombre, db.func.count(Nominacion.id))
        .join(Nominacion, Nominacion.valor_id == Valor.id)
        .filter(Nominacion.ciclo_id == ciclo.id)
        .group_by(Valor.nombre)
        .order_by(db.func.count(Nominacion.id).desc())
        .limit(10)
        .all()
    )

    por_dia = (
        db.session.query(db.func.date(Nominacion.fecha), db.func.count(Nominacion.id))
        .filter(Nominacion.ciclo_id == ciclo.id)
        .group_by(db.func.date(Nominacion.fecha))
        .order_by(db.func.date(Nominacion.fecha))
        .all()
    )

    return jsonify({
        "totales": {
            "maestros": total_maestros,
            "alumnos": total_alumnos,
            "nominaciones": total_nom
        },
        "por_tipo": por_tipo_dict,
        "por_valor": [{"valor": v, "n": n} for v, n in por_valor],
        "por_dia": [{"fecha": str(f), "n": n} for f, n in por_dia],
    })


# ===============================
# 🗓️ CALENDARIO DE EVENTOS
# ===============================
from datetime import datetime
from models import EventoAsamblea, PlantillaInvitacion, Bloque, CicloEscolar

@admin_bp.route('/calendario', methods=['GET', 'POST'])
@login_required
@admin_required
def calendario_eventos():
    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo_activo:
        flash("⚠️ No hay un ciclo escolar activo.", "warning")
        return redirect(url_for('admin_bp.gestionar_ciclos'))

    bloques = (
        Bloque.query
        .filter_by(ciclo_id=ciclo_activo.id)
        .order_by(cast(func.substr(Bloque.nombre, 8), Integer).asc())
        .all()
    )
    plantillas = PlantillaInvitacion.query.filter_by(ciclo_id=ciclo_activo.id, activa=True).all()

    if request.method == 'POST':
        bloque_id = request.form.get('bloque_id')
        nombre_mes = request.form.get('nombre_mes')
        mes_ordinal = request.form.get('mes_ordinal', type=int)
        fecha_evento = request.form.get('fecha_evento')
        fecha_cierre = request.form.get('fecha_cierre_nominaciones')
        plantilla_id = request.form.get('plantilla_id') or None

        if not all([bloque_id, nombre_mes, mes_ordinal, fecha_evento, fecha_cierre]):
            flash("⚠️ Todos los campos son obligatorios.", "warning")
            return redirect(url_for('admin_bp.calendario_eventos'))

        # 🧩 Validación: la fecha de cierre debe ser menor a la fecha del evento
        try:
            fecha_evento_dt = datetime.fromisoformat(fecha_evento)
            fecha_cierre_dt = datetime.fromisoformat(fecha_cierre)

            if fecha_cierre_dt >= fecha_evento_dt:
                flash("⚠️ La fecha de cierre de nominaciones debe ser anterior a la fecha del evento.", "danger")
                return redirect(url_for('admin_bp.calendario_eventos'))

        except Exception as e:
            flash(f"❌ Error al procesar las fechas: {str(e)}", "danger")
            return redirect(url_for('admin_bp.calendario_eventos'))

        # ✅ Crear evento si las fechas son válidas
        evento = EventoAsamblea(
            ciclo_id=ciclo_activo.id,
            bloque_id=bloque_id,
            nombre_mes=nombre_mes.strip().capitalize(),
            mes_ordinal=mes_ordinal,
            fecha_evento=fecha_evento,
            fecha_cierre_nominaciones=fecha_cierre,
            plantilla_id=plantilla_id
        )
        db.session.add(evento)
        db.session.commit()
        flash("✅ Evento agregado exitosamente.", "success")
        return redirect(url_for('admin_bp.calendario_eventos'))

    # Mostrar todos los eventos existentes del ciclo actual
    eventos = (
        EventoAsamblea.query
        .filter_by(ciclo_id=ciclo_activo.id)
        .order_by(EventoAsamblea.mes_ordinal.asc(), EventoAsamblea.bloque_id.asc())
        .all()
    )
    return render_template(
        'admin_calendario.html',
        ciclo=ciclo_activo,
        bloques=bloques,
        plantillas=plantillas,
        eventos=eventos,
        datetime=datetime
    )

# ===============================
# 🗑️ ELIMINAR EVENTO DE ASAMBLEA (AJAX)
# ===============================
@admin_bp.route('/calendario/eliminar/<int:evento_id>', methods=['DELETE'])
@login_required
@admin_required
def eliminar_evento(evento_id):
    evento = EventoAsamblea.query.get(evento_id)
    if not evento:
        return jsonify({'success': False, 'error': 'Evento no encontrado'}), 404

    db.session.delete(evento)
    db.session.commit()
    return jsonify({'success': True})
# ===============================
# 🔄 ACTIVAR / DESACTIVAR EVENTO
# ===============================
@admin_bp.route('/calendario/toggle/<int:evento_id>', methods=['POST'])
@login_required
@admin_required
def toggle_evento(evento_id):
    evento = EventoAsamblea.query.get(evento_id)
    if not evento:
        return jsonify({'success': False, 'error': 'Evento no encontrado'}), 404

    evento.activo = not evento.activo
    db.session.commit()
    return jsonify({'success': True, 'activo': evento.activo})

# ===============================
# 📊 MATRIZ MENSUAL (Datos JSON)
# ===============================
from models import Alumno, Nominacion, EventoAsamblea, Maestro, Valor, Bloque, CicloEscolar

@admin_bp.route('/matriz_data', methods=['GET'])
@login_required
@admin_required
def matriz_data():
    """Devuelve los datos estructurados de la matriz mensual."""
    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo_activo:
        return jsonify({"error": "No hay ciclo activo."}), 400

    bloque_id = request.args.get('bloque_id', type=int)
    grado = request.args.get('grado')
    grupo = request.args.get('grupo')

    # 1️⃣ Obtener eventos del ciclo ordenados por mes
    eventos = EventoAsamblea.query.filter_by(ciclo_id=ciclo_activo.id).order_by(EventoAsamblea.mes_ordinal.asc()).all()
    meses_data = [
        {
            "id": e.id,
            "nombre": e.nombre_mes,
            "mes_ordinal": e.mes_ordinal,
            "cierre": e.fecha_cierre_nominaciones.strftime("%Y-%m-%d %H:%M")
        }
        for e in eventos
    ]

    # 2️⃣ Filtrar alumnos según bloque / grado / grupo
    q = Alumno.query.filter_by(ciclo_id=ciclo_activo.id)
    if bloque_id:
        q = q.filter(Alumno.bloque_id == bloque_id)
    if grado:
        q = q.filter(Alumno.grado == grado)
    if grupo:
        q = q.filter(Alumno.grupo == grupo.upper())

    alumnos = q.order_by(Alumno.grado, Alumno.grupo, Alumno.nombre).all()

    # 3️⃣ Armar estructura de alumnos con sus nominaciones por mes
    resultado = []
    for alumno in alumnos:
        valores_por_mes = {str(e.mes_ordinal): [] for e in eventos}
        nominaciones = Nominacion.query.filter_by(alumno_id=alumno.id, ciclo_id=ciclo_activo.id).all()

        for n in nominaciones:
            if n.evento and n.evento.mes_ordinal:
                valores_por_mes[str(n.evento.mes_ordinal)].append({
                    "valor": n.valor.nombre if n.valor else "",
                    "maestro": n.maestro.nombre if n.maestro else "",
                    "fecha": n.fecha.strftime("%Y-%m-%d") if n.fecha else ""
                })

        sin_nominacion = all(len(v) == 0 for v in valores_por_mes.values())

        resultado.append({
            "id": alumno.id,
            "nombre": alumno.nombre,
            "grado": alumno.grado,
            "grupo": alumno.grupo,
            "valores_por_mes": valores_por_mes,
            "sin_nominacion": sin_nominacion
        })

    return jsonify({
        "ciclo": ciclo_activo.nombre,
        "bloque": Bloque.query.get(bloque_id).nombre if bloque_id else None,
        "meses": meses_data,
        "alumnos": resultado
    })

@admin_bp.route('/bloques_json')
@login_required
@admin_required
def bloques_json():
    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo_activo:
        return jsonify([])
    bloques = Bloque.query.filter_by(ciclo_id=ciclo_activo.id).order_by(Bloque.nombre.asc()).all()
    return jsonify([{"id": b.id, "nombre": b.nombre} for b in bloques])

@admin_bp.route('/matriz')
@login_required
@admin_required
def matriz_vista():
    """Vista HTML de la matriz mensual."""
    return render_template('admin_matriz.html')

# ===============================
# 👨‍🏫 PANEL DEL MAESTRO (vista previa tipo matriz)
# ===============================
from flask import jsonify
from models import Alumno, Nominacion, EventoAsamblea, CicloEscolar, Maestro, Valor

@nom.route('/panel_nominaciones')
@login_required
def panel_nominaciones():
    """Vista HTML del panel de nominaciones (para maestros)."""
    if current_user.rol != 'profesor':
        flash("🚫 Solo los profesores pueden acceder al panel de nominaciones.", "danger")
        return redirect(url_for('nom.principal'))
    return render_template('panel_maestro.html')


@nom.route('/panel_nominaciones_data')
@login_required
def panel_nominaciones_data():
    """Devuelve los datos para la matriz del maestro actual (con encabezado informativo)."""
    if current_user.rol != 'profesor':
        return jsonify({"error": "No autorizado."}), 403

    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo_activo:
        return jsonify({"error": "No hay ciclo activo."}), 400

    maestro = Maestro.query.filter_by(correo=current_user.email, ciclo_id=ciclo_activo.id).first()
    if not maestro:
        return jsonify({"error": "No se encontró tu registro de maestro."}), 404

    # --- Evento activo del ciclo (si hay)
    evento_activo = (
        EventoAsamblea.query
        .filter_by(ciclo_id=ciclo_activo.id, activo=True)
        .order_by(EventoAsamblea.fecha_evento.asc())
        .first()
    )

    # --- Lista de meses del ciclo
    eventos = EventoAsamblea.query.filter_by(ciclo_id=ciclo_activo.id).order_by(EventoAsamblea.mes_ordinal.asc()).all()
    meses = [{"id": e.id, "nombre": e.nombre_mes, "mes_ordinal": e.mes_ordinal} for e in eventos]

    # --- Detectar bloque actual a partir de los alumnos del ciclo
    primer_alumno = Alumno.query.filter_by(ciclo_id=ciclo_activo.id).first()
    bloque_id = primer_alumno.bloque_id if primer_alumno else None
    bloque_nombre = None
    grado_actual = primer_alumno.grado if primer_alumno else None
    grupo_actual = primer_alumno.grupo if primer_alumno else None

    if bloque_id:
        from models import Bloque
        bloque_obj = Bloque.query.get(bloque_id)
        bloque_nombre = bloque_obj.nombre if bloque_obj else None

    # --- Filtrar alumnos del mismo bloque / grado / grupo
    alumnos = (
        Alumno.query
        .filter_by(ciclo_id=ciclo_activo.id, bloque_id=bloque_id, grado=grado_actual, grupo=grupo_actual)
        .order_by(Alumno.nombre.asc())
        .all()
    )

    resultado = []
    for alumno in alumnos:
        valores_por_mes = {str(e.mes_ordinal): [] for e in eventos}
        nominaciones = Nominacion.query.filter_by(alumno_id=alumno.id, ciclo_id=ciclo_activo.id).all()

        for n in nominaciones:
            if n.evento and n.evento.mes_ordinal:
                valores_por_mes[str(n.evento.mes_ordinal)].append({
                    "valor": n.valor.nombre if n.valor else "",
                    "maestro": n.maestro.nombre if n.maestro else "",
                    "fecha": n.fecha.strftime("%Y-%m-%d") if n.fecha else ""
                })

        sin_nominacion = all(len(v) == 0 for v in valores_por_mes.values())

        resultado.append({
            "id": alumno.id,
            "nombre": alumno.nombre,
            "grado": alumno.grado,
            "grupo": alumno.grupo,
            "valores_por_mes": valores_por_mes,
            "sin_nominacion": sin_nominacion
        })

    # Detectar bloque automáticamente a partir de los alumnos cargados
    bloque_nombre = None
    if alumnos:
        primer_alumno = alumnos[0]
        if primer_alumno.bloque_id:
            from models import Bloque
            bloque_obj = Bloque.query.get(primer_alumno.bloque_id)
            bloque_nombre = bloque_obj.nombre if bloque_obj else None

    return jsonify({
        "bloque": bloque_nombre or "Sin bloque",
        "grado": alumnos[0].grado if alumnos else "—",
        "grupo": alumnos[0].grupo if alumnos else "—",
        "ciclo": ciclo_activo.nombre,
        "evento_activo": evento_activo.nombre_mes if evento_activo else None,
        "cierre_evento": evento_activo.fecha_cierre_nominaciones.strftime("%d/%m/%Y %H:%M") if evento_activo else None,
        "meses": meses,
        "alumnos": resultado
    })


# ===============================
# 🧾 NOMINACIÓN INDIVIDUAL DE ALUMNO (con lógica de EXCELENCIA)
# ===============================
from datetime import datetime

@nom.route('/nominacion_alumno/<int:alumno_id>', methods=['GET', 'POST'])
@login_required
def nominar_alumno_individual(alumno_id):
    cerrar_eventos_vencidos()
    """Vista individual donde el maestro puede nominar a un alumno."""

    # 1️⃣ Verificar rol del usuario
    if current_user.rol != 'profesor':
        flash("🚫 Solo los profesores pueden registrar nominaciones.", "danger")
        return redirect(url_for('nom.principal'))

    # 2️⃣ Verificar ciclo activo
    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo_activo:
        flash("⚠️ No hay ciclo activo disponible.", "warning")
        return redirect(url_for('nom.panel_nominaciones'))

    # 3️⃣ Maestro actual
    maestro = Maestro.query.filter_by(correo=current_user.email, ciclo_id=ciclo_activo.id).first()
    alumno = Alumno.query.get_or_404(alumno_id)

    # 4️⃣ Valores activos del ciclo
    valores = Valor.query.filter_by(ciclo_id=ciclo_activo.id, activo=True).all()

    # 5️⃣ Detectar evento activo del bloque
    evento_abierto = (
        EventoAsamblea.query
        .filter_by(ciclo_id=ciclo_activo.id, bloque_id=alumno.bloque_id, activo=True)
        .order_by(EventoAsamblea.fecha_evento.asc())
        .first()
    )

    # 🚫 Validar que el evento esté abierto
    if not evento_abierto:
        flash("🚫 Las nominaciones han cerrado para este bloque.", "danger")
        return redirect(url_for(
            'nom.matriz_grupo_maestro',
            bloque_id=alumno.bloque_id,
            grado=alumno.grado,
            grupo=alumno.grupo
        ))

    # 6️⃣ Filtrar nominaciones del evento actual (mismo mes)
    nominaciones_previas = (
        Nominacion.query
        .filter_by(alumno_id=alumno.id, ciclo_id=ciclo_activo.id, evento_id=evento_abierto.id)
        .all()
    )

    # 7️⃣ Obtener valores ya usados
    valores_asignados = [n.valor_id for n in nominaciones_previas]
    valores_disponibles = [v for v in valores if v.id not in valores_asignados]

    # 8️⃣ Procesar formulario
    if request.method == 'POST':
        valor_id = request.form.get('valor_id')
        comentario = request.form.get('comentario', '').strip()

        if not valor_id:
            flash("⚠️ Debes seleccionar un valor.", "warning")
            return redirect(request.url)

        # 🚫 Bloquear si el alumno ya tiene EXCELENCIA
        tiene_excelencia = (
            Nominacion.query
            .join(Valor)
            .filter(
                Nominacion.alumno_id == alumno.id,
                Nominacion.ciclo_id == ciclo_activo.id,
                Valor.nombre == "EXCELENCIA"
            )
            .first()
        )
        if tiene_excelencia:
            flash("🏆 Este alumno ya alcanzó el valor máximo EXCELENCIA y no puede recibir más nominaciones.", "warning")
            return redirect(url_for(
                'nom.matriz_grupo_maestro',
                bloque_id=alumno.bloque_id,
                grado=alumno.grado,
                grupo=alumno.grupo
            ))

        # ✅ Registrar la nueva nominación
        nueva_nom = Nominacion(
            alumno_id=alumno.id,
            maestro_id=maestro.id,
            valor_id=valor_id,
            ciclo_id=ciclo_activo.id,
            comentario=comentario,
            evento_id=evento_abierto.id,
            tipo='alumno',
            fecha=datetime.utcnow()
        )
        db.session.add(nueva_nom)
        db.session.commit()

        # 🧠 Verificar si alcanza EXCELENCIA
        promovido = actualizar_a_excelencia(alumno.id, ciclo_activo.id)

        if promovido:
            flash(f"🏅 {alumno.nombre} ha alcanzado el valor EXCELENCIA por acumular 3 nominaciones.", "success")
        else:
            flash(f"✅ Nominación registrada para {alumno.nombre}.", "success")

        return redirect(url_for(
            'nom.matriz_grupo_maestro',
            bloque_id=alumno.bloque_id,
            grado=alumno.grado,
            grupo=alumno.grupo
        ))

    # 9️⃣ Renderizar plantilla
    return render_template(
        'nominacion_individual.html',
        alumno=alumno,
        valores_disponibles=valores_disponibles,
        valores_asignados=nominaciones_previas
    )

# ===============================
# 🧭 PANEL PRINCIPAL DEL PROFESOR
# ===============================
@nom.route('/panel_profesor')
@login_required
def panel_profesor():
    """Panel principal con accesos rápidos para el profesor."""
    if current_user.rol != 'profesor':
        flash("🚫 Solo los profesores pueden acceder a este panel.", "danger")
        return redirect(url_for('nom.principal'))

    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()
    return render_template('panel_profesor.html', ciclo=ciclo_activo)


# =======================================
# 👨‍🏫 VISTA "MIS GRUPOS" (profesor optimizada y ordenada)
@nom.route('/mis_grupos')
@login_required
def mis_grupos():
    """Muestra los grupos y alumnos con orden correcto, límites y bloqueo por excelencia."""
    if current_user.rol != 'profesor':
        flash("🚫 Solo los profesores pueden acceder a esta vista.", "danger")
        return redirect(url_for('nom.principal'))

    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo_activo:
        flash("⚠️ No hay ciclo escolar activo.", "warning")
        return redirect(url_for('nom.panel_profesor'))

    maestro = Maestro.query.filter_by(correo=current_user.email, ciclo_id=ciclo_activo.id).first()
    if not maestro:
        flash("⚠️ No se encontró tu registro como maestro activo.", "warning")
        return redirect(url_for('nom.panel_profesor'))

    # 🔹 Traer todo en una sola consulta (optimizado)
    alumnos = (
        Alumno.query
        .filter_by(ciclo_id=ciclo_activo.id)
        .order_by(Alumno.grado, Alumno.grupo, Alumno.nombre)
        .all()
    )

    # 🔹 Traer TODOS los eventos activos del ciclo y mapearlos por bloque
    eventos_activos = (
        EventoAsamblea.query
        .filter_by(ciclo_id=ciclo_activo.id, activo=True)
        .all()
    )

    evento_por_bloque = {}
    for e in eventos_activos:
        # Si hay varios activos en el mismo bloque, nos quedamos con el más reciente por fecha_evento
        actual = evento_por_bloque.get(e.bloque_id)
        if not actual or e.fecha_evento > actual.fecha_evento:
            evento_por_bloque[e.bloque_id] = e

    # Cargar todas las nominaciones del ciclo de una sola vez
    nominaciones_ciclo = Nominacion.query.filter_by(ciclo_id=ciclo_activo.id).all()
    nominaciones_por_alumno = {}
    for n in nominaciones_ciclo:
        nominaciones_por_alumno.setdefault(n.alumno_id, []).append(n)

    grupos = {}

    # Construir estructura base de grupos
    for alumno in alumnos:
        clave = f"{alumno.grado}°{alumno.grupo}"
        if clave not in grupos:
            grupos[clave] = []

        nominaciones = nominaciones_por_alumno.get(alumno.id, [])

        # 🟡 Verificar excelencia
        tiene_excelencia = any(
            n.valor and getattr(n.valor, "nombre", "").upper() == "EXCELENCIA"
            for n in nominaciones
        )

        # 🟢 Verificar nominación del mes actual PERO por BLOQUE del alumno
        tiene_nominaciones_mes = False
        evento_bloque = evento_por_bloque.get(alumno.bloque_id)
        if evento_bloque:
            # Nominaciones ligadas al evento de ese bloque
            tiene_nominaciones_mes = any(
                n.evento_id == evento_bloque.id for n in nominaciones
            )

        grupos[clave].append({
            "id": alumno.id,
            "nombre": alumno.nombre,
            "tiene_excelencia": tiene_excelencia,
            "tiene_nominaciones_mes": tiene_nominaciones_mes
        })

    # ---------- Parser robusto del token después de "°" (para ordenar) ----------
    import re

    def parse_grupo_key(nombre_grupo: str):
        """
        nombre_grupo ejemplo: '01 - PRIMERO°K1 A', '03 - TERCERO°PP1 B', '05 - QUINTO°P5 C', '01 - PRIMERO°SEC1 A'
        Devuelve una tupla (orden_seccion, grado_num, letra_grupo)
        para ordenar correctamente Kinder → Preprimaria → Primaria → Secundaria.
        """
        t = nombre_grupo.upper().strip()
        m = re.search(r'°\s*([A-Z]+)(\d+)\b', t)
        if not m:
            return (99, 99, 'Z')
        seccion_code = m.group(1)
        grado_num = int(m.group(2))
        m2 = re.search(r'\s([A-Z])$', t)
        letra = m2.group(1) if m2 else ''
        orden_seccion = {'K': 1, 'PP': 2, 'P': 3, 'SEC': 4}.get(seccion_code, 9)
        return (orden_seccion, grado_num, letra)

    grupos_ordenados = dict(sorted(grupos.items(), key=lambda kv: parse_grupo_key(kv[0])))

    # 🔹 Detectar límite automáticamente
    def obtener_limite(nombre_grupo):
        nombre = nombre_grupo.upper()
        if "SEC" in nombre or "SECUND" in nombre or "BLOQUE 4" in nombre:
            return 11
        return 8

    # 🔹 Numerar alumnos dentro de cada grupo
    for grupo_nombre, lista_alumnos in grupos_ordenados.items():
        for i, alumno in enumerate(lista_alumnos, start=1):
            alumno["numero"] = i

    return render_template(
        "mis_grupos.html",
        grupos=grupos_ordenados,
        ciclo=ciclo_activo,
        obtener_limite=obtener_limite
    )


@nom.route('/seleccionar_bloque')
@login_required
def seleccionar_bloque():
    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()

    # 🔹 Ordenar por número si el nombre contiene un número (Bloque 1, Bloque 2, etc.)
    bloques = (
        Bloque.query
        .filter_by(ciclo_id=ciclo_activo.id)
        .order_by(db.cast(db.func.substr(Bloque.nombre, 8), db.Integer))  # Ordena por el número después de "Bloque "
        .all()
    )

    return render_template('seleccionar_bloque.html', bloques=bloques, ciclo=ciclo_activo)

# -------------------------------
# 🔹 Seleccionar grado dentro del bloque
# -------------------------------
@nom.route('/bloque/<int:bloque_id>/grados')
@login_required
def seleccionar_grado(bloque_id):
    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()
    bloque = Bloque.query.get_or_404(bloque_id)

    # 🔍 Obtener los grados únicos dentro del bloque
    grados = (
        db.session.query(Alumno.grado)
        .filter_by(bloque_id=bloque.id, ciclo_id=ciclo_activo.id)
        .distinct()
        .order_by(Alumno.grado)
        .all()
    )
    grados = [g[0] for g in grados]  # Convierte [(1,), (2,), (3,)] → [1,2,3]

    return render_template('seleccionar_grado.html', bloque=bloque, grados=grados, ciclo=ciclo_activo)


# -------------------------------
# 🔹 Panel de gestión de usuarios
# -------------------------------


@admin_bp.route('/admin/maestros')
@login_required
@admin_required
def admin_maestros():
    return redirect(url_for('admin_bp.alumnos_maestros'))  # si ya existe el template o ruta
    # o bien:
    # return render_template('admin_maestros.html')

@admin_bp.route('/usuarios/<rol>')
@login_required
@admin_required
def admin_usuarios_por_rol(rol):
    """Muestra usuarios filtrados por rol (admin, profesor, estudiante)"""
    roles_validos = ['admin', 'profesor', 'estudiante']
    if rol not in roles_validos:
        flash("Rol no válido.", "danger")
        return redirect(url_for('nom.principal'))

    # Filtrar usuarios según el rol
    usuarios = Usuario.query.filter_by(rol=rol).all()
    return render_template('admin_usuarios.html', usuarios=usuarios, rol_actual=rol)

# =========================================================
# 🗑️ ELIMINAR USUARIO (ADMIN, PROFESOR O ESTUDIANTE)
# =========================================================
@admin_bp.route('/usuarios/eliminar/<int:id>', methods=['POST'])
@login_required
@admin_required
def eliminar_usuario(id):
    usuario = Usuario.query.get_or_404(id)

    # 🚫 Evita eliminar el admin principal (puedes añadir más correos si quieres protegerlos)
    if usuario.email.lower() in ["admin@cela.edu.mx", "admin@colegio.edu.mx"]:
        return jsonify({
            "success": False,
            "message": "⚠️ No puedes eliminar el administrador principal."
        }), 400

    db.session.delete(usuario)
    db.session.commit()

    return jsonify({
        "success": True,
        "message": f"🗑️ Usuario '{usuario.nombre}' eliminado correctamente."
    }), 200


# ============================================================
# 🔹 GESTIÓN DE MAESTROS
# ============================================================
@admin_bp.route('/maestros', methods=['GET'])
@login_required
@admin_required
def gestionar_maestros():
    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo_activo:
        flash("⚠️ No hay ciclo escolar activo.", "warning")
        return redirect(url_for('admin_bp.gestionar_ciclos'))
    
    maestros = Maestro.query.filter_by(ciclo_id=ciclo_activo.id, activo=True).all()
    return render_template('admin_maestros.html', maestros=maestros)


# ------------------------------------------------------------
# 🟢 Crear maestro manualmente (con rol profesor y contraseña)
# ------------------------------------------------------------
@admin_bp.route('/maestros/nuevo', methods=['GET', 'POST'])
@login_required
@admin_required
def nuevo_maestro():
    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo_activo:
        flash("⚠️ No hay un ciclo activo para registrar maestros.", "warning")
        return redirect(url_for('admin_bp.maestros_ciclo'))

    if request.method == 'POST':
        nombre = request.form.get('nombre', '').strip()
        correo = request.form.get('correo', '').strip().lower()
        password = request.form.get('password', '').strip()

        # Validar campos obligatorios
        if not nombre or not correo or not password:
            flash("⚠️ Todos los campos son obligatorios.", "warning")
            return redirect(url_for('admin_bp.nuevo_maestro'))

        # Verificar si ya existe un usuario con ese correo
        usuario_existente = Usuario.query.filter_by(email=correo).first()
        if usuario_existente:
            flash("❌ Ya existe un usuario con ese correo electrónico.", "danger")
            return redirect(url_for('admin_bp.maestros_ciclo'))

        # Crear usuario con rol profesor
        nuevo_usuario = Usuario(nombre=nombre, email=correo, rol='profesor')
        nuevo_usuario.set_password(password)
        db.session.add(nuevo_usuario)

        # Crear registro en tabla maestros
        nuevo_maestro = Maestro(
            nombre=nombre,
            correo=correo,
            ciclo_id=ciclo_activo.id,
            activo=True
        )
        db.session.add(nuevo_maestro)

        db.session.commit()
        flash(f"✅ Maestro '{nombre}' creado correctamente con rol 'profesor'.", "success")
        return redirect(url_for('admin_bp.maestros_ciclo'))

    return render_template('crear_maestro.html')


@admin_bp.route('/alumnos/crear', methods=['GET', 'POST'])
@login_required
@admin_required
def crear_alumno_manual():
    ciclo = CicloEscolar.query.filter_by(activo=True).first()
    bloques = Bloque.query.filter_by(ciclo_id=ciclo.id).all() if ciclo else []

    if request.method == 'POST':
        nombre = request.form.get('nombre')
        grado = request.form.get('grado')
        grupo = request.form.get('grupo')
        nivel = request.form.get('nivel')
        bloque_id = request.form.get('bloque_id')

        if not all([nombre, grado, grupo, nivel, bloque_id]):
            flash("⚠️ Todos los campos son obligatorios.", "warning")
            return redirect(url_for('admin_bp.crear_alumno_manual'))

        nuevo = Alumno(
            nombre=nombre,
            grado=grado,
            grupo=grupo,
            nivel=nivel,
            bloque_id=bloque_id,
            ciclo_id=ciclo.id
        )
        db.session.add(nuevo)
        db.session.commit()
        flash(f"✅ Alumno {nombre} registrado correctamente.", "success")
        return redirect(url_for('admin_bp.maestros_ciclo'))

    return render_template('crear_alumno.html', ciclo=ciclo, bloques=bloques)


# Muestra los grupos disponibles dentro de un grado
@nom.route('/bloque/<int:bloque_id>/grado/<grado>')
@login_required
def grupos_por_grado(bloque_id, grado):
    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo_activo:
        flash("⚠️ No hay ciclo activo.", "warning")
        return redirect(url_for('nom.principal'))

    grupos = (
        db.session.query(Alumno.grupo)
        .filter_by(ciclo_id=ciclo_activo.id, bloque_id=bloque_id, grado=grado)
        .distinct()
        .order_by(Alumno.grupo.asc())
        .all()
    )
    grupos = [g[0] for g in grupos]
    bloque = Bloque.query.get_or_404(bloque_id)
    return render_template('grupos_por_grado.html', bloque=bloque, grado=grado, grupos=grupos)

# Vista temporal para mostrar alumnos de ese grupo
@nom.route('/bloque/<int:bloque_id>/grado/<grado>/grupo/<grupo>')
@login_required
def lista_alumnos_grupo(bloque_id, grado, grupo):
    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()
    alumnos = (
        Alumno.query
        .filter_by(ciclo_id=ciclo_activo.id, bloque_id=bloque_id, grado=grado, grupo=grupo)
        .order_by(Alumno.nombre.asc())
        .all()
    )
    bloque = Bloque.query.get_or_404(bloque_id)
    return render_template('lista_alumnos.html', alumnos=alumnos, bloque=bloque, grado=grado, grupo=grupo)

# ===========================
# 🔹 MATRIZ DE NOMINACIÓN POR GRUPO (MAESTRO)
# ===========================
# 🔹 MATRIZ DE NOMINACIÓN POR GRUPO (MAESTRO)
# ===========================
from datetime import datetime

@nom.route('/bloque/<int:bloque_id>/grado/<grado>/grupo/<grupo>/nominaciones')
@login_required
def matriz_grupo_maestro(bloque_id, grado, grupo):
    if current_user.rol != 'profesor':
        flash("🚫 Solo los profesores pueden acceder a esta vista.", "danger")
        return redirect(url_for('nom.principal'))

    hoy = datetime.now().date()
    ciclo = CicloEscolar.query.filter_by(activo=True).first()
    bloque = Bloque.query.get_or_404(bloque_id)

    # 🔹 Buscar evento vigente (activo y dentro de fechas válidas)
    evento_abierto = (
        EventoAsamblea.query
        .filter(
            EventoAsamblea.ciclo_id == ciclo.id,
            EventoAsamblea.bloque_id == bloque.id,
            EventoAsamblea.activo == True,
            EventoAsamblea.fecha_evento >= hoy  # aún no ha pasado
        )
        .order_by(EventoAsamblea.fecha_evento.asc())
        .first()
    )

    # Si no hay futuros, usar el último pasado (el más reciente)
    if not evento_abierto:
        evento_abierto = (
            EventoAsamblea.query
            .filter_by(ciclo_id=ciclo.id, bloque_id=bloque.id)
            .order_by(EventoAsamblea.fecha_evento.desc())
            .first()
        )

    mes_actual = evento_abierto.mes_ordinal if evento_abierto else None

    # 🔹 Cargar todos los eventos del ciclo del bloque
    eventos = (
        EventoAsamblea.query
        .filter_by(ciclo_id=ciclo.id, bloque_id=bloque.id)
        .order_by(EventoAsamblea.mes_ordinal.asc())
        .all()
    )

    # 🔹 Agrupar por mes
    eventos_por_mes = {}
    for e in eventos:
        if e.mes_ordinal not in eventos_por_mes:
            eventos_por_mes[e.mes_ordinal] = e

    # 🔹 Cargar alumnos del grupo
    alumnos = (
        Alumno.query
        .filter_by(ciclo_id=ciclo.id, bloque_id=bloque.id, grado=grado, grupo=grupo)
        .order_by(Alumno.nombre.asc())
        .all()
    )

    data = []
    for a in alumnos:
        nominaciones = (
            Nominacion.query
            .filter_by(alumno_id=a.id, ciclo_id=ciclo.id, tipo='alumno')
            .join(EventoAsamblea, Nominacion.evento_id == EventoAsamblea.id)
            .add_entity(EventoAsamblea)
            .all()
        )

        valores_por_mes = {str(e.mes_ordinal): [] for e in eventos_por_mes.values()}

        for n, evento in nominaciones:
            if evento and evento.mes_ordinal:
                valores_por_mes[str(evento.mes_ordinal)].append({
                    "valor": n.valor.nombre if n.valor else "",
                    "maestro": n.maestro.nombre if n.maestro else "",
                    "fecha": n.fecha.strftime("%Y-%m-%d") if n.fecha else ""
                })

        tiene_nominaciones = any(len(v) > 0 for v in valores_por_mes.values())
        sin_nominacion_total = not tiene_nominaciones

        data.append({
            "alumno": a,
            "valores": valores_por_mes,
            "sin_nominacion_total": sin_nominacion_total
        })

    return render_template(
        "matriz_grupo_maestro.html",
        ciclo=ciclo,
        bloque=bloque,
        grado=grado,
        grupo=grupo,
        eventos=list(eventos_por_mes.values()),
        alumnos=data,
        mes_actual=mes_actual,
        evento_abierto=evento_abierto
    )


@nom.route('/mis_nominaciones', methods=['GET'])
@login_required
def mis_nominaciones():
    if current_user.rol != 'profesor':
        flash("🚫 Solo los profesores pueden acceder a esta vista.", "danger")
        return redirect(url_for('nom.principal'))

    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo_activo:
        flash("⚠️ No hay ciclo escolar activo.", "warning")
        return redirect(url_for('nom.panel_profesor'))

    maestro = Maestro.query.filter_by(correo=current_user.email, ciclo_id=ciclo_activo.id).first()
    if not maestro:
        flash("⚠️ No se encontró tu registro como maestro en el ciclo activo.", "warning")
        return redirect(url_for('nom.panel_profesor'))

    # Todos los eventos del ciclo
    eventos_ciclo = (
        EventoAsamblea.query
        .filter_by(ciclo_id=ciclo_activo.id)
        .order_by(EventoAsamblea.mes_ordinal, EventoAsamblea.bloque_id)
        .all()
    )

    # Un evento representativo por mes para el selector (solo para nombre y orden)
    eventos_unicos = []
    meses_vistos = set()
    for e in eventos_ciclo:
        if e.nombre_mes not in meses_vistos:
            eventos_unicos.append(e)
            meses_vistos.add(e.nombre_mes)

    # Mapa: mes -> ¿hay al menos un evento abierto en ese mes?
    ahora = datetime.utcnow()
    mes_abierto_map = {}
    for e in eventos_ciclo:
        abierto = (e.activo is True) and (e.fecha_cierre_nominaciones and e.fecha_cierre_nominaciones > ahora)
        mes_abierto_map[e.nombre_mes] = mes_abierto_map.get(e.nombre_mes, False) or abierto

    # Mes seleccionado (querystring) o, si no hay, el primer mes que tenga algún evento abierto; si no, el más reciente
    mes_seleccionado = request.args.get('mes')
    if not mes_seleccionado:
        abierto_ordenado = [e.nombre_mes for e in eventos_ciclo
                            if (e.activo is True and e.fecha_cierre_nominaciones and e.fecha_cierre_nominaciones > ahora)]
        if abierto_ordenado:
            mes_seleccionado = abierto_ordenado[0]
        elif eventos_ciclo:
            mes_seleccionado = sorted({e.nombre_mes for e in eventos_ciclo},
                                    key=lambda m: next(x.mes_ordinal for x in eventos_ciclo if x.nombre_mes == m))[-1]
        else:
            mes_seleccionado = None

    # Nominaciones del maestro del mes seleccionado
    if mes_seleccionado:
        nominaciones = (
            Nominacion.query
            .filter_by(maestro_id=maestro.id, ciclo_id=ciclo_activo.id)
            .join(EventoAsamblea)
            .filter(EventoAsamblea.nombre_mes == mes_seleccionado)
            .order_by(Nominacion.fecha.desc())
            .all()
        )
    else:
        nominaciones = []

    # Eventos ABiertos de ese mes (pueden ser varios bloques)
    eventos_abiertos_mes = (
        EventoAsamblea.query
        .filter(
            EventoAsamblea.ciclo_id == ciclo_activo.id,
            EventoAsamblea.nombre_mes == mes_seleccionado,
            EventoAsamblea.activo.is_(True),
            EventoAsamblea.fecha_cierre_nominaciones > ahora
        )
        .all()
        if mes_seleccionado else []
    )
    open_event_ids = {e.id for e in eventos_abiertos_mes}
    evento_abierto = eventos_abiertos_mes[0] if eventos_abiertos_mes else None  # para el banner verde/rojo

    # Valores disponibles
    valores = Valor.query.filter_by(ciclo_id=ciclo_activo.id, activo=True).all()
    valores_json = [{"id": v.id, "nombre": v.nombre} for v in valores]

    # Mapa de valores permitidos por nominación
    allowed_map = {}
    for n in nominaciones:
        if n.tipo == "alumno":
            usados = {nom.valor_id for nom in Nominacion.query.filter_by(
                ciclo_id=ciclo_activo.id, tipo='alumno', alumno_id=n.alumno_id
            )}
        else:
            usados = {nom.valor_id for nom in Nominacion.query.filter_by(
                ciclo_id=ciclo_activo.id, tipo='personal',
                maestro_id=n.maestro_id, maestro_nominado_id=n.maestro_nominado_id
            )}
        disponibles = [v for v in valores if v.id not in usados or v.id == n.valor_id]
        allowed_map[n.id] = [{"id": v.id, "nombre": v.nombre} for v in disponibles]

    return render_template(
        'mis_nominaciones.html',
        maestro=maestro,
        nominaciones=nominaciones,
        ciclo=ciclo_activo,
        valores_json=valores_json,
        allowed_valores_map_json=allowed_map,
        eventos=eventos_unicos,
        mes_seleccionado=mes_seleccionado,
        evento_abierto=evento_abierto,
        eventos_abiertos_mes=eventos_abiertos_mes,
        open_event_ids=open_event_ids,
        mes_abierto_map=mes_abierto_map
    )

from sqlalchemy import cast, Integer, func
# ===========================
# 🔹 Nueva vista visual de bloques
# ===========================
@nom.route('/admin/bloques_vista', methods=['GET'])
@login_required
def admin_bloques_vista():
    if current_user.rol != 'admin':
        flash("🚫 Solo los administradores pueden ver esta sección.", "danger")
        return redirect(url_for('nom.principal'))

    ciclo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo:
        flash("⚠️ No hay ciclo escolar activo.", "warning")
        return redirect(url_for('nom.principal'))

    bloques = (
    Bloque.query
    .filter_by(ciclo_id=ciclo.id)
    .order_by(cast(func.substr(Bloque.nombre, 8), Integer).asc())
    .all()
    )

    data = []
    for b in bloques:
        grados_raw = db.session.query(Alumno.grado).filter(
            Alumno.ciclo_id == ciclo.id,
            Alumno.bloque_id == b.id
        ).distinct().all()
        grados = [g[0] for g in grados_raw]

        grados_info = []
        for g in grados:
            grupos_raw = db.session.query(Alumno.grupo).filter(
                Alumno.ciclo_id == ciclo.id,
                Alumno.bloque_id == b.id,
                Alumno.grado == g
            ).distinct().all()
            grupos = [x[0] for x in grupos_raw]
            grados_info.append({"grado": g, "grupos": grupos})

        data.append({"bloque": b, "grados": grados_info})

    return render_template('admin_bloques_vista.html', ciclo=ciclo, data=data)

import io
import pandas as pd
from flask import send_file, jsonify

@admin_bp.route('/maestros/plantilla', methods=['GET'])
@login_required
@admin_required
def descargar_plantilla_maestros():
    # Define columnas esperadas
    columnas = ['Nombre', 'Email', "contraseña"]
    df = pd.DataFrame(columns=columnas)

    # Crear archivo Excel en memoria
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Plantilla')
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='Plantilla_Maestros.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # ------------------------------------------------------------
# 🔄 Actualizar orden del bloque
# ------------------------------------------------------------
@admin_bp.route('/bloques/<int:id>/orden', methods=['POST'])
@login_required
def actualizar_orden_bloque(id):
    if current_user.rol != 'admin':
        flash("🚫 Solo los administradores pueden realizar esta acción.", "danger")
        return redirect(url_for('nom.principal'))

    bloque = Bloque.query.get_or_404(id)
    nuevo_orden = request.form.get('orden', type=int)

    if nuevo_orden is None:
        flash("⚠️ El valor de orden no es válido.", "warning")
    else:
        bloque.orden = nuevo_orden
        db.session.commit()
        flash(f"✅ Orden del bloque '{bloque.nombre}' actualizado a {nuevo_orden}.", "success")

    return redirect(url_for('admin_bp.bloques_ciclo'))


# ------------------------------------------------------------
# 🗑️ Eliminar bloque (solo si no tiene alumnos)
# ------------------------------------------------------------
@admin_bp.route('/bloques/<int:id>/eliminar', methods=['POST'])
@login_required
def eliminar_bloque(id):
    if current_user.rol != 'admin':
        flash("🚫 Solo los administradores pueden realizar esta acción.", "danger")
        return redirect(url_for('nom.principal'))

    bloque = Bloque.query.get_or_404(id)
    alumnos_asociados = Alumno.query.filter_by(bloque_id=bloque.id).count()

    if alumnos_asociados > 0:
        flash(f"⚠️ No se puede eliminar el bloque '{bloque.nombre}' porque tiene alumnos asociados.", "warning")
    else:
        db.session.delete(bloque)
        db.session.commit()
        flash(f"🗑️ Bloque '{bloque.nombre}' eliminado correctamente.", "success")

    return redirect(url_for('admin_bp.bloques_ciclo'))

# -------------------------------
# 🧩 Extensiones permitidas para importar alumnos
# -------------------------------
ALLOWED_EXT_ALUMNOS = {"xlsx", "csv"}

def _allowed_file_alumnos(filename: str) -> bool:
    """Verifica si el archivo tiene una extensión válida (.xlsx o .csv)"""
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXT_ALUMNOS

# -------------------------------
# 🔹 Eliminar nominación de personal
# -------------------------------
@nom.route('/nominaciones/personal/eliminar/<int:id>', methods=['GET'])
@login_required
def eliminar_nominacion_personal(id):
    nominacion = Nominacion.query.get_or_404(id)

    # Validar que sea del maestro actual
    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()
    maestro = Maestro.query.filter_by(correo=current_user.email, ciclo_id=ciclo_activo.id).first()

    if not maestro or nominacion.maestro_id != maestro.id:
        flash("🚫 No tienes permiso para eliminar esta nominación.", "danger")
        return redirect(url_for('nom.nominar_personal'))

    db.session.delete(nominacion)
    db.session.commit()
    flash("🗑️ Nominación eliminada correctamente.", "success")
    return redirect(url_for('nom.nominar_personal'))

# -------------------------------
# 🔹 Editar nominación de personal
# -------------------------------
@nom.route('/nominaciones/personal/editar/<int:id>', methods=['POST'])
@login_required
def editar_nominacion_personal(id):
    nominacion = Nominacion.query.get_or_404(id)

    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()
    maestro = Maestro.query.filter_by(correo=current_user.email, ciclo_id=ciclo_activo.id).first()

    if not maestro or nominacion.maestro_id != maestro.id:
        flash("🚫 No puedes editar esta nominación.", "danger")
        return redirect(url_for('nom.nominar_personal'))

    nuevo_valor = request.form.get('valor_id')
    nuevo_comentario = request.form.get('comentario', '').strip()

    if nuevo_valor:
        nominacion.valor_id = nuevo_valor
    nominacion.comentario = nuevo_comentario if nuevo_comentario else None
    nominacion.fecha = datetime.utcnow()

    db.session.commit()
    
    
    flash("✏️ Nominación actualizada correctamente.", "success")
    return redirect(url_for('nom.nominar_personal'))

# ============================================
# ✏️ Editar nominación (blindada + validación de duplicados + recalcula EXCELENCIA)
# ============================================
@nom.route('/nominaciones/editar/<int:id>', methods=['POST'])
@login_required
def editar_nominacion(id):
    from flask import jsonify

    nominacion = Nominacion.query.get_or_404(id)
    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()

    # 🔐 Validar rol y propiedad
    if current_user.rol != 'profesor':
        return jsonify({"status": "error", "message": "Solo los maestros pueden editar nominaciones."}), 403

    maestro = Maestro.query.filter_by(correo=current_user.email, ciclo_id=ciclo_activo.id).first()
    if not maestro or nominacion.maestro_id != maestro.id:
        return jsonify({"status": "error", "message": "No tienes permiso para editar esta nominación."}), 403

    # 🚫 Verificar si el evento ya cerró
    if not nominacion.evento or not nominacion.evento.esta_abierto:
        return jsonify({
            "status": "error",
            "message": f"Las nominaciones del mes '{nominacion.evento.nombre_mes}' están cerradas."
        }), 403

    # 🧩 Obtener datos enviados
    data = request.get_json()
    try:
        valor_id = int(data.get('valor_id'))
    except (TypeError, ValueError):
        return jsonify({"status": "error", "message": "Valor inválido."}), 400

    comentario = (data.get('comentario') or '').strip()
    ciclo_id = nominacion.ciclo_id

    # 🔒 Validar duplicados según tipo (mantiene tu lógica original)
    if nominacion.tipo == 'alumno' and nominacion.alumno_id:
        existe = Nominacion.query.filter(
            Nominacion.ciclo_id == ciclo_id,
            Nominacion.tipo == 'alumno',
            Nominacion.alumno_id == nominacion.alumno_id,
            Nominacion.valor_id == valor_id,
            Nominacion.id != nominacion.id
        ).first()
        if existe:
            return jsonify({
                "status": "error",
                "message": "Ese alumno ya tiene asignado ese valor en este ciclo."
            }), 400

    if nominacion.tipo == 'personal' and nominacion.maestro_nominado_id:
        existe = Nominacion.query.filter(
            Nominacion.ciclo_id == ciclo_id,
            Nominacion.tipo == 'personal',
            Nominacion.maestro_id == nominacion.maestro_id,
            Nominacion.maestro_nominado_id == nominacion.maestro_nominado_id,
            Nominacion.valor_id == valor_id,
            Nominacion.id != nominacion.id
        ).first()
        if existe:
            return jsonify({
                "status": "error",
                "message": "Ya registraste ese mismo valor para este maestro en este ciclo."
            }), 400

    # ✅ Guardar cambios base
    nominacion.valor_id = valor_id
    nominacion.comentario = comentario

    # 🟡 Reapegar etiqueta de control si aplica
    if "[EXCELENCIA-VISUAL]" not in (nominacion.comentario or ""):
        nominacion.comentario = (nominacion.comentario or "").strip() + " [EXCELENCIA-VISUAL]"

    db.session.commit()

    # 🔁 Intentar recalcular la nominación EXCELENCIA asociada
    try:
        recalcular_comentario_excelencia(nominacion.alumno_id, nominacion.ciclo_id)
    except Exception as e:
        print("⚠️ Error recalculando EXCELENCIA:", e)

    return jsonify({
        "status": "success",
        "message": "✅ Nominación actualizada correctamente y EXCELENCIA sincronizada."
    })



# ============================================
# 🗑️ Eliminar nominación (blindada)
# ============================================
@nom.route('/nominaciones/eliminar/<int:id>', methods=['POST'])
@login_required
def eliminar_nominacion(id):
    nominacion = Nominacion.query.get_or_404(id)
    ciclo_activo = CicloEscolar.query.filter_by(activo=True).first()

    # 🔐 Validar rol y propiedad
    if current_user.rol != 'profesor':
        flash("🚫 Solo los maestros pueden eliminar nominaciones.", "danger")
        return redirect(url_for('nom.mis_nominaciones'))

    maestro = Maestro.query.filter_by(correo=current_user.email, ciclo_id=ciclo_activo.id).first()
    if not maestro or nominacion.maestro_id != maestro.id:
        flash("⚠️ No tienes permiso para eliminar esta nominación.", "warning")
        return redirect(url_for('nom.mis_nominaciones'))

    # 🚫 Verificar si el evento está cerrado
    if not nominacion.evento or not nominacion.evento.esta_abierto:
        flash(f"⚠️ Las nominaciones del mes '{nominacion.evento.nombre_mes}' están cerradas. No puedes eliminarla.", "warning")
        return redirect(url_for('nom.mis_nominaciones'))

    # ✅ Eliminar si todo está correcto
    db.session.delete(nominacion)
    db.session.commit()
    verificar_reversion_excelencia(nominacion.alumno_id, nominacion.ciclo_id)
    flash("🗑️ Nominación eliminada correctamente.", "success")
    return redirect(url_for('nom.mis_nominaciones'))

@admin_bp.route('/dashboard/nominaciones_maestro', methods=['GET'])
@login_required
@admin_required
def nominaciones_por_maestro_y_mes():
    ciclo = CicloEscolar.query.filter_by(activo=True).first()
    maestro_id = request.args.get('maestro_id', type=int)
    mes_ordinal = request.args.get('mes', type=int)

    if not ciclo or not maestro_id or not mes_ordinal:
        return jsonify({"error": "Parámetros incompletos"}), 400

    nominaciones = (
        Nominacion.query
        .join(EventoAsamblea)
        .filter(
            Nominacion.ciclo_id == ciclo.id,
            Nominacion.maestro_id == maestro_id,
            EventoAsamblea.mes_ordinal == mes_ordinal
        )
        .order_by(Nominacion.fecha.asc())
        .all()
    )

    data = []
    for n in nominaciones:
        # 🔹 Determinar bloque solo si es alumno
        if n.tipo == "alumno" and n.alumno:
            bloque = n.alumno.bloque.nombre
        else:
            bloque = "—"

        data.append({
            "id": n.id,
            "tipo": n.tipo,
            "nominado": (
                n.alumno.nombre if n.alumno
                else n.maestro_nominado.nombre if n.maestro_nominado
                else ""
            ),
            "valor": n.valor.nombre if n.valor else "",
            "fecha": n.fecha.strftime("%Y-%m-%d") if n.fecha else "",
            "comentario": n.comentario or "",
            "evento": n.evento.nombre_mes if n.evento else "",
            "bloque": bloque  # 🔹 Nuevo campo agregado
        })

    return jsonify(data)



# ======================================================
# == Generar DOCX y ZIP EN MEMORIA (versión universal segura, con filtro EXCELENCIA)
# ======================================================
@nom.route('/admin/dashboard/generar_invitaciones_stream')
@login_required
@admin_required
def generar_invitaciones_stream():
    import io, zipfile, time, os, gc
    from flask import send_file, request, make_response
    from docxtpl import DocxTemplate
    from sqlalchemy.orm import joinedload
    from models import Nominacion, CicloEscolar

    ids = request.args.get("ids", "")
    if not ids:
        return "No se especificaron IDs", 400

    ids = [int(i) for i in ids.split(",")]

    ciclo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo:
        return "No hay ciclo activo", 400

    nominaciones = (
        Nominacion.query
        .options(
            joinedload(Nominacion.maestro),
            joinedload(Nominacion.alumno),
            joinedload(Nominacion.maestro_nominado),
            joinedload(Nominacion.valor),
            joinedload(Nominacion.evento)
        )
        .filter(Nominacion.id.in_(ids))
        .all()
    )

    if not nominaciones:
        return "No se encontraron nominaciones.", 404

    # =====================================================
    # 🔹 Filtrar duplicados: si un alumno tiene EXCELENCIA,
    # solo se exporta esa nominación (no las otras)
    # =====================================================
    nominaciones_filtradas = []
    procesados = set()

    for n in nominaciones:
        if n.alumno_id in procesados:
            continue

        tiene_excelencia = any(
            x.alumno_id == n.alumno_id and x.valor and x.valor.nombre.upper() == "EXCELENCIA"
            for x in nominaciones
        )

        if tiene_excelencia:
            excelencia = next(
                (x for x in nominaciones if x.alumno_id == n.alumno_id and x.valor and x.valor.nombre.upper() == "EXCELENCIA"),
                None
            )
            if excelencia:
                nominaciones_filtradas.append(excelencia)
            procesados.add(n.alumno_id)
        else:
            nominaciones_filtradas.append(n)
            procesados.add(n.alumno_id)

    nominaciones = nominaciones_filtradas

    # =====================================================
    # 🔹 Generar ZIP totalmente en memoria
    # =====================================================
    mem_zip = io.BytesIO()
    with zipfile.ZipFile(mem_zip, "w", zipfile.ZIP_DEFLATED) as zf:
        for n in nominaciones:
            try:
                tipo = n.tipo or "alumno"
                plantilla = (
                    "formato_asamblea.docx"
                    if tipo == "alumno"
                    else "invitacion colaborador 1.docx"
                )
                doc = DocxTemplate(os.path.join("docx_templates", plantilla))

                # =====================================================
                # 🔹 Ajuste del comentario para EXCELENCIA
                # =====================================================
                comentario_final = n.comentario or ""
                if n.valor and n.valor.nombre.upper() == "EXCELENCIA":
                    # Eliminar tags visuales
                    comentario_final = comentario_final.replace("[EXCELENCIA-VISUAL]", "").replace("  ", " ").strip()

                    # Reemplazar texto genérico
                    if "Valores obtenidos:" in comentario_final:
                        comentario_final = comentario_final.replace("Valores obtenidos:", "Por sus valores de")
                    if "Comentarios:" in comentario_final:
                        comentario_final = comentario_final.replace("Comentarios:", "— Comentarios de los maestros:")

                    # Limpieza adicional (quitar números tipo 1 | 2 | 3)
                    comentario_final = comentario_final.replace("1 |", "").replace("2 |", "").replace("3 |", "").strip()
                    comentario_final = comentario_final.replace("|", " ").replace("  ", " ").strip()

                # =====================================================
                # 🔹 Contexto del documento
                # =====================================================
                context = {
                    "quien_nomina": n.maestro.nombre if n.maestro else "",
                    "nominado": (
                        n.alumno.nombre if tipo == "alumno"
                        else n.maestro_nominado.nombre if n.maestro_nominado else ""
                    ),
                    "valor": n.valor.nombre if n.valor else "",
                    "fecha_evento": n.evento.fecha_evento.strftime("%d/%m/%Y") if n.evento else "",
                    "texto_adicional": comentario_final,
                }

                doc.render(context)

                temp = io.BytesIO()
                doc.save(temp)
                temp.seek(0)

                nombre_nominado = n.alumno.nombre if tipo == "alumno" else (
                    n.maestro_nominado.nombre if n.maestro_nominado else ""
                )
                filename = f"{nombre_nominado.replace(' ', '_')}_{tipo}_{n.valor.nombre if n.valor else 'SinValor'}.docx"
                zf.writestr(filename, temp.read())
                temp.close()
                gc.collect()
            except Exception as e:
                print(f"⚠️ Error generando invitación {n.id}: {e}")

    mem_zip.seek(0)
    filename_zip = f"invitaciones_{ciclo.nombre}_{time.strftime('%Y%m%d_%H%M')}.zip"

    # ✅ Generamos respuesta explícita y segura
    response = make_response(mem_zip.read())
    response.headers["Content-Type"] = "application/zip"
    response.headers["Content-Disposition"] = f"attachment; filename={filename_zip}"
    response.headers["Content-Length"] = str(mem_zip.tell())
    response.headers["Cache-Control"] = "no-store"
    mem_zip.close()

    return response

# ==========================
# 🌟 MURO PÚBLICO DE NOMINADOS
# ==========================

from sqlalchemy.orm import joinedload
from collections import defaultdict

@nom.route('/muro_publico')
def muro_publico():
    """Vista pública de nominados por mes, bloque y tipo."""
    ciclo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo:
        return render_template("public_muro.html", error="No hay ciclo activo actualmente.")

    # Parámetros de filtro
    mes = request.args.get("mes", type=int)
    tipo = request.args.get("tipo", "alumno")
    bloque_id = request.args.get("bloque", type=int)

    # Obtener eventos del ciclo activo y agrupar meses únicos
    eventos_query = EventoAsamblea.query.filter_by(ciclo_id=ciclo.id).order_by(EventoAsamblea.mes_ordinal).all()

    # 🔹 Extraer solo meses únicos (por nombre y ordinal)
    eventos_unicos = []
    vistos = set()
    for e in eventos_query:
        if e.nombre_mes not in vistos:
            eventos_unicos.append(e)
            vistos.add(e.nombre_mes)

    # Si no se especifica mes, tomar el último evento activo
    if not mes and eventos_unicos:
        mes = max(e.mes_ordinal for e in eventos_unicos)

    # 📦 Obtener lista de bloques (para el filtro)
    bloques = Bloque.query.filter_by(ciclo_id=ciclo.id).order_by(Bloque.orden).all()

    # Filtrar nominaciones por ciclo, mes, tipo y bloque (si aplica)
    query = Nominacion.query.options(
        joinedload(Nominacion.valor),
        joinedload(Nominacion.alumno),
        joinedload(Nominacion.maestro_nominado),
        joinedload(Nominacion.evento)
    ).filter(
        Nominacion.ciclo_id == ciclo.id,
        Nominacion.tipo == tipo,
        Nominacion.evento.has(EventoAsamblea.mes_ordinal == mes)
    )

    if bloque_id:
        query = query.join(Alumno).filter(Alumno.bloque_id == bloque_id)

    nominaciones_raw = query.all()

    # 🧩 Agrupar por persona (alumno o maestro nominado)
    agrupadas = defaultdict(lambda: {"nombre": "", "bloque": "", "valores": set(), "evento": None})

    for n in nominaciones_raw:
        if tipo == "alumno" and n.alumno:
            clave = n.alumno.id
            agrupadas[clave]["nombre"] = n.alumno.nombre
            agrupadas[clave]["bloque"] = n.alumno.bloque.nombre if n.alumno.bloque else ""
        elif tipo == "personal" and n.maestro_nominado:
            clave = n.maestro_nominado.id
            agrupadas[clave]["nombre"] = n.maestro_nominado.nombre
            agrupadas[clave]["bloque"] = "-"
        else:
            continue

        agrupadas[clave]["valores"].add(n.valor.nombre)
        agrupadas[clave]["evento"] = n.evento.nombre_mes

    # 🧮 Convertir a lista y ordenar por cantidad de valores (mayor a menor)
    nominaciones = sorted(
        agrupadas.values(),
        key=lambda x: len(x["valores"]),
        reverse=True
    )

    return render_template(
        "public_muro.html",
        ciclo=ciclo,
        eventos=eventos_unicos,
        nominaciones=nominaciones,
        mes=mes,
        tipo=tipo,
        bloque_id=bloque_id,
        bloques=bloques
    )
@nom.route('/inicio_rapido')
@login_required
def inicio_rapido():
    """Redirige al panel correcto según el rol del usuario"""
    if current_user.rol == 'admin':
        return redirect(url_for('nom.panel_admin'))
    elif current_user.rol == 'profesor':
        return redirect(url_for('nom.panel_profesor'))  # ✅ nombre correcto
    elif current_user.rol == 'alumno':
        return redirect(url_for('nom.panel_alumno'))  # si llegas a tenerlo
    else:
        flash("Rol no reconocido. Contacte al administrador.", "warning")
        return redirect(url_for('nom.logout'))
    
# ======================================================
# 📦 Generar invitaciones por bloque (solo exporta EXCELENCIA si existe)
# ======================================================
# ======================================================
# 📦 Generar invitaciones por bloque (solo exporta EXCELENCIA si existe)
# ======================================================
@nom.route('/admin/dashboard/generar_invitaciones_bloque_unico')
@login_required
@admin_required
def generar_invitaciones_bloque_unico():
    import io, zipfile, time, os, re, gc
    from flask import make_response, request
    from sqlalchemy.orm import joinedload
    from docxtpl import DocxTemplate
    from models import Nominacion, Alumno, Bloque, CicloEscolar, EventoAsamblea

    # ----------------------------------------
    # 🔹 Leer parámetros
    # ----------------------------------------
    bloque_id = request.args.get("bloque_id", type=int)
    mes_nombre = request.args.get("mes", type=str)

    if not bloque_id:
        return "No se especificó bloque.", 400

    ciclo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo:
        return "No hay ciclo activo.", 400

    bloque = Bloque.query.get(bloque_id)
    if not bloque:
        return "Bloque no encontrado.", 404

    # ----------------------------------------
    # 🔹 Si viene mes, buscar evento exacto
    # ----------------------------------------
    evento_filtrado = None
    if mes_nombre:
        evento_filtrado = EventoAsamblea.query.filter_by(
            bloque_id=bloque_id,
            nombre_mes=mes_nombre,
            ciclo_id=ciclo.id
        ).first()

        if not evento_filtrado:
            return f"No hay evento de {mes_nombre} para este bloque.", 404

    # =====================================================
    # 🔹 Traer nominaciones del bloque filtradas por mes
    # =====================================================
    query = (
        Nominacion.query
        .options(
            joinedload(Nominacion.maestro),
            joinedload(Nominacion.alumno),
            joinedload(Nominacion.maestro_nominado),
            joinedload(Nominacion.valor),
            joinedload(Nominacion.evento),
        )
        .join(Alumno, Alumno.id == Nominacion.alumno_id)
        .filter(
            Nominacion.ciclo_id == ciclo.id,
            Alumno.bloque_id == bloque.id
        )
        .order_by(Nominacion.fecha.asc())
    )

    # 🔹 Si hay mes → filtrar por evento específico
    if evento_filtrado:
        query = query.filter(Nominacion.evento_id == evento_filtrado.id)

    nominaciones = query.all()

    if not nominaciones:
        return f"No hay nominaciones para el mes {mes_nombre} en {bloque.nombre}.", 404

    # =====================================================
    # 🔹 Filtrar duplicados con EXCELENCIA
    # =====================================================
    nominaciones_filtradas = []
    procesados = set()

    for n in nominaciones:
        if n.alumno_id in procesados:
            continue

        tiene_excelencia = any(
            x.alumno_id == n.alumno_id and x.valor and x.valor.nombre.upper() == "EXCELENCIA"
            for x in nominaciones
        )

        if tiene_excelencia:
            excelencia = next(
                (x for x in nominaciones
                    if x.alumno_id == n.alumno_id and x.valor and x.valor.nombre.upper() == "EXCELENCIA"),
                None
            )
            if excelencia:
                nominaciones_filtradas.append(excelencia)
            procesados.add(n.alumno_id)
        else:
            nominaciones_filtradas.append(n)
            procesados.add(n.alumno_id)

    nominaciones = nominaciones_filtradas

    # =====================================================
    # 🔹 Generar ZIP totalmente en memoria
    # =====================================================
    def slug(s):
        s = (s or "").strip()
        s = re.sub(r"[^\w\-\.]+", "_", s, flags=re.UNICODE)
        return s[:80] or "archivo"

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for n in nominaciones:
            try:
                tipo = (n.tipo or "alumno").strip().lower()
                plantilla = (
                    "formato_asamblea.docx" if tipo == "alumno"
                    else "invitacion colaborador 1.docx"
                )
                plantilla_path = os.path.join("docx_templates", plantilla)

                doc = DocxTemplate(plantilla_path)
                nominado = (
                    n.alumno.nombre if tipo == "alumno"
                    else n.maestro_nominado.nombre if n.maestro_nominado else ""
                )

                # -------- reconstrucción de comentario excelencia --------
                comentario_final = n.comentario or ""
                # =====================================================
                # 🧠 Reconstrucción elegante de comentarios para EXCELENCIA
                # =====================================================
                if n.valor and n.valor.nombre.upper() == "EXCELENCIA":

                    # 1. Obtener nominaciones previas (las que formaron la excelencia)
                    nominaciones_previas = (
                        Nominacion.query
                        .filter(
                            Nominacion.alumno_id == n.alumno_id,
                            Nominacion.ciclo_id == n.ciclo_id,
                            Nominacion.valor_id != n.valor_id  # excluir EXCELENCIA
                        )
                        .order_by(Nominacion.fecha.asc())
                        .all()
                    )

                    # 2. Lista de valores previos (para ponerlos en la primera línea)
                    valores_previos = [
                        nom.valor.nombre
                        for nom in nominaciones_previas
                        if nom.valor and nom.valor.nombre.upper() != "EXCELENCIA"
                    ]

                    valores_texto = ", ".join(valores_previos)

                    # 3. Agrupar comentarios por maestro
                    comentarios_por_maestro = {}

                    for nom in nominaciones_previas:
                        maestro_nombre = nom.maestro.nombre if nom.maestro else "Maestro desconocido"
                        comentario = (nom.comentario or "").replace("[EXCELENCIA-VISUAL]", "").strip()

                        if comentario:
                            comentarios_por_maestro.setdefault(maestro_nombre, []).append(comentario)

                    # 4. Construir texto final
                    comentario_final = f"Por sus valores de {valores_texto}.\n"
                    comentario_final += "— Comentarios de los maestros:\n"

                    for maestro_nombre, comentarios in comentarios_por_maestro.items():
                        comentario_final += f"{maestro_nombre}:\n"
                        for c in comentarios:
                            comentario_final += f"• {c}\n"
                        comentario_final += ""  # espacio entre maestros

                # -------- Render --------
                context = {
                    "quien_nomina": n.maestro.nombre if n.maestro else "",
                    "nominado": nominado,
                    "valor": n.valor.nombre if n.valor else "",
                    "fecha_evento": n.evento.fecha_evento.strftime("%d/%m/%Y") if n.evento else "",
                    "texto_adicional": comentario_final,
                }

                doc_io = io.BytesIO()
                doc.render(context)
                doc.save(doc_io)
                doc_io.seek(0)

                filename = f"{slug(nominado)}_{slug(tipo)}_{slug(n.valor.nombre if n.valor else 'SinValor')}.docx"
                zf.writestr(filename, doc_io.read())
                doc_io.close()
                gc.collect()
            except Exception as e:
                print(f"⚠️ Error generando invitación NominacionID={n.id}: {e}")

    zip_buffer.seek(0)
    filename_zip = f"Invitaciones_{slug(bloque.nombre)}_{slug(mes_nombre)}_{time.strftime('%Y%m%d_%H%M')}.zip"

    response = make_response(zip_buffer.read())
    response.headers["Content-Type"] = "application/zip"
    response.headers["Content-Disposition"] = f"attachment; filename={filename_zip}"
    response.headers["Cache-Control"] = "no-store"

    zip_buffer.close()
    print(f"✅ Exportado bloque {bloque.nombre} mes {mes_nombre} ({len(nominaciones)} invitaciones generadas)")

    return response

# ======================================================
# == Exportar concentrado general de nominaciones (Excel)
# ======================================================
# ======================================================
# == Exportar concentrado general de nominaciones (Excel)
# ======================================================
@nom.route('/admin/dashboard/exportar_concentrado_excel')
@login_required
@admin_required
def exportar_concentrado_excel():
    import io
    import pandas as pd
    from flask import make_response
    from datetime import datetime
    from models import Nominacion, CicloEscolar, Alumno, Valor, EventoAsamblea, Bloque
    from sqlalchemy.orm import joinedload
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

    ciclo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo:
        return "No hay ciclo activo.", 400

    # 🔹 Cargar todas las nominaciones (con relaciones)
    nominaciones = (
        Nominacion.query
        .options(
            joinedload(Nominacion.maestro),
            joinedload(Nominacion.alumno).joinedload(Alumno.bloque),
            joinedload(Nominacion.maestro_nominado),
            joinedload(Nominacion.valor),
            joinedload(Nominacion.evento),
        )
        .filter(Nominacion.ciclo_id == ciclo.id)
        .filter(~Nominacion.comentario.contains("[EXCELENCIA-VISUAL]"))  # ⛔ excluir visuales
        .order_by(Nominacion.fecha.asc())
        .all()
    )

    if not nominaciones:
        return "No hay nominaciones registradas.", 404

    # 🔹 Construcción de los datos para DataFrames
    data_alumnos, data_personal = [], []

    for n in nominaciones:
        tipo = n.tipo or "alumno"
        evento = (
            getattr(n.evento, "nombre", None)
            or getattr(n.evento, "nombre_mes", None)
            or getattr(n.evento, "titulo", None)
            or getattr(n.evento, "titulo_evento", None)
            or "—"
        )

        base = {
            "Quién nomina": n.maestro.nombre if n.maestro else "",
            "Tipo": tipo.capitalize(),
            "Nominado": (
                n.alumno.nombre if tipo == "alumno"
                else n.maestro_nominado.nombre if n.maestro_nominado else ""
            ),
            "Bloque": n.alumno.bloque.nombre if n.alumno and n.alumno.bloque else "—",
            "Grado": n.alumno.grado if n.alumno and hasattr(n.alumno, "grado") else "—",
            "Grupo": n.alumno.grupo if n.alumno and hasattr(n.alumno, "grupo") else "—",
            "Valor": n.valor.nombre if n.valor else "",
            "Comentario": n.comentario or "",
            "Evento": evento,
            "Fecha": n.fecha.strftime("%d/%m/%Y") if n.fecha else "",
        }

        if tipo == "alumno":
            data_alumnos.append(base)
        else:
            data_personal.append(base)

    # Crear DataFrames
    df_alumnos = pd.DataFrame(data_alumnos)
    df_personal = pd.DataFrame(data_personal)

    # 🧱 Generar Excel
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # --- HOJA 1: ALUMNOS ---
        if not df_alumnos.empty:
            df_alumnos = df_alumnos.sort_values(by=["Grupo", "Bloque", "Grado"])
            df_alumnos.to_excel(writer, index=False, sheet_name="Alumnos")

            ws = writer.book["Alumnos"]
            header_fill = PatternFill(start_color="7B0000", end_color="7B0000", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            header_align = Alignment(horizontal="center", vertical="center")
            border = Border(
                left=Side(style="thin", color="999999"),
                right=Side(style="thin", color="999999"),
                top=Side(style="thin", color="999999"),
                bottom=Side(style="thin", color="999999"),
            )

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
                cell.border = border

            # 🎨 Alternar colores por grupo
            colores = ["FFF8E1", "FFFFFF"]
            ultimo_grupo = None
            color_idx = 0
            for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
                grupo = str(ws[f"F{i}"].value or "").strip()  # columna F = Grupo
                if grupo != ultimo_grupo:
                    color_idx = 1 - color_idx
                    ultimo_grupo = grupo
                fill = PatternFill(start_color=colores[color_idx], end_color=colores[color_idx], fill_type="solid")
                for c in row:
                    c.fill = fill
                    c.border = border

            ws.freeze_panes = "A2"
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 45)

        # --- HOJA 2: PERSONAL ---
        if not df_personal.empty:
            df_personal = df_personal.sort_values(by=["Fecha"])
            df_personal.to_excel(writer, index=False, sheet_name="Personal")

            ws2 = writer.book["Personal"]
            for cell in ws2[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_align
                cell.border = border

            ws2.freeze_panes = "A2"
            for col in ws2.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                ws2.column_dimensions[col[0].column_letter].width = min(max_len + 3, 45)

    # 📦 Enviar respuesta
    output.seek(0)
    filename = f"Concentrado_Nominaciones_{ciclo.nombre}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    response = make_response(output.read())
    response.headers["Content-Disposition"] = f"attachment; filename={filename}"
    response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    response.headers["Cache-Control"] = "no-store"
    output.close()

    return response

# ======================================================
# 📦 Generar invitaciones para PROFESORES (colaboradores)
# ======================================================
# ======================================================
# 📦 Generar invitaciones para PROFESORES (colaboradores)
# ======================================================
@nom.route('/admin/dashboard/generar_invitaciones_profesores')
@login_required
@admin_required
def generar_invitaciones_profesores():
    import io, zipfile, time, os, re, gc
    from flask import make_response, request
    from sqlalchemy.orm import joinedload
    from docxtpl import DocxTemplate
    from models import Nominacion, CicloEscolar

    def slug(s):
        s = (s or "").strip()
        s = re.sub(r"[^\w\-\.]+", "_", s, flags=re.UNICODE)
        return s[:80] or "archivo"

    ciclo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo:
        return "No hay ciclo activo", 400

    # 🔹 Obtener mes desde el dashboard (Enero, Noviembre, etc.)
    mes_nombre = request.args.get("mes")
    if not mes_nombre:
        return "No se especificó el mes.", 400

    # 🔹 Filtrar por mes → necesitamos el mes_ordinal correspondiente
    #    Ej: Enero → 2, Noviembre → 1
    from models import EventoAsamblea
    evento_mes = EventoAsamblea.query.filter_by(
        ciclo_id=ciclo.id,
        nombre_mes=mes_nombre
    ).first()

    if not evento_mes:
        return f"No existe un evento para el mes {mes_nombre}.", 404

    mes_ordinal = evento_mes.mes_ordinal

    # =====================================================
    # 🔹 Solo profesores (tipo personal) del MES indicado
    # =====================================================
    nominaciones = (
        Nominacion.query
        .options(
            joinedload(Nominacion.maestro),
            joinedload(Nominacion.maestro_nominado),
            joinedload(Nominacion.valor),
            joinedload(Nominacion.evento),
        )
        .filter(
            Nominacion.ciclo_id == ciclo.id,
            Nominacion.tipo == "personal",
            Nominacion.evento_id == evento_mes.id    # 👈 FILTRAMOS SOLO ESE MES
        )
        .order_by(Nominacion.fecha.asc())
        .all()
    )

    if not nominaciones:
        return f"No hay nominaciones para profesores en {mes_nombre}.", 404

    # =====================================================
    # 🔹 Generación del ZIP (idéntico al tuyo)
    # =====================================================
    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:

        for n in nominaciones:
            try:
                plantilla_path = os.path.join("docx_templates", "invitacion colaborador 1.docx")
                if not os.path.exists(plantilla_path):
                    return f"No se encontró la plantilla DOCX: {plantilla_path}", 500

                doc = DocxTemplate(plantilla_path)

                nominado = n.maestro_nominado.nombre if n.maestro_nominado else ""
                valor = n.valor.nombre if n.valor else ""
                comentario_final = (n.comentario or "").strip()

                fecha_evento = (
                    n.evento.fecha_evento.strftime("%d/%m/%Y")
                    if n.evento and n.evento.fecha_evento else ""
                )

                contexto = {
                    "quien_nomina": n.maestro.nombre if n.maestro else "",
                    "nominado": nominado,
                    "valor": valor,
                    "fecha_evento": fecha_evento,
                    "texto_adicional": comentario_final,
                }

                doc_io = io.BytesIO()
                doc.render(contexto)
                doc.save(doc_io)
                doc_io.seek(0)

                filename = f"{slug(nominado)}_{slug(valor)}.docx"
                zf.writestr(filename, doc_io.read())
                doc_io.close()
                gc.collect()

            except Exception as e:
                print(f"⚠️ Error generando invitación profesor (ID={n.id}): {e}")

    zip_buffer.seek(0)
    filename_zip = f"Invitaciones_Profesores_{mes_nombre}_{slug(ciclo.nombre)}_{time.strftime('%Y%m%d_%H%M')}.zip"

    response = make_response(zip_buffer.read())
    response.headers["Content-Type"] = "application/zip"
    response.headers["Content-Disposition"] = f"attachment; filename={filename_zip}"
    response.headers["Cache-Control"] = "no-store"
    return response

# ======================================================
# 🧹 Gestor de nominaciones (NUEVA PÁGINA)
# URL: /admin/gestor_nominaciones
# ======================================================
from flask import jsonify, request
from extensions import db
from models import Nominacion, Alumno, Maestro, Bloque, Valor, EventoAsamblea, CicloEscolar

@admin_bp.route('/gestor_nominaciones')
@login_required
@admin_required
def admin_gestor_nominaciones():
    """
    Renderiza la NUEVA página de gestión de nominaciones.
    """
    ciclo = CicloEscolar.query.filter_by(activo=True).first()
    return render_template('admin_gestor_nominaciones.html', ciclo=ciclo)


@admin_bp.route('/gestor_nominaciones/data')
@login_required
@admin_required
def admin_gestor_nominaciones_data():
    """
    Devuelve todas las nominaciones del ciclo activo en JSON para DataTables.
    """
    ciclo = CicloEscolar.query.filter_by(activo=True).first()
    if not ciclo:
        return jsonify([])

    nominaciones = (
        Nominacion.query
        .filter_by(ciclo_id=ciclo.id)
        .options(
            db.joinedload(Nominacion.alumno).joinedload(Alumno.bloque),
            db.joinedload(Nominacion.maestro),
            db.joinedload(Nominacion.maestro_nominado),
            db.joinedload(Nominacion.valor),
            db.joinedload(Nominacion.evento),
        )
        .order_by(Nominacion.fecha.desc())
        .all()
    )

    data = []
    for n in nominaciones:
        tipo = n.tipo or ""
        if tipo == "personal" and n.maestro_nominado:
            nominado = n.maestro_nominado.nombre
        else:
            nominado = n.alumno.nombre if n.alumno else "—"

        bloque = n.alumno.bloque.nombre if (n.alumno and n.alumno.bloque) else "—"
        valor = n.valor.nombre if n.valor else "—"
        maestro = n.maestro.nombre if n.maestro else "—"
        fecha = n.fecha.strftime("%d/%m/%Y") if n.fecha else "—"

        if n.evento:
            evento_label = f"{n.evento.nombre_mes} ({n.evento.fecha_evento.strftime('%d/%m/%Y')})"
        else:
            evento_label = "—"

        data.append({
            "id": n.id,
            "fecha": fecha,
            "mes": n.evento.nombre_mes if n.evento else "—",
            "tipo": tipo,
            "nominado": nominado,
            "maestro": maestro,
            "bloque": bloque,
            "valor": valor,
            "comentario": (n.comentario or "").strip(),
            "evento": evento_label
        })
    return jsonify(data)


@admin_bp.route('/gestor_nominaciones/eliminar', methods=['POST'])
@login_required
@admin_required
def admin_gestor_nominaciones_eliminar():
    """
    Elimina una o varias nominaciones mediante IDs enviados en JSON.
    """
    data = request.get_json(silent=True) or {}
    ids = data.get("ids", [])

    if not ids:
        return jsonify({"success": False, "message": "No se recibieron IDs."}), 400

    try:
        Nominacion.query.filter(Nominacion.id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()
        return jsonify({"success": True, "eliminadas": len(ids)})
    except Exception as e:
        db.session.rollback()
        print("❌ Error al eliminar nominaciones:", e)
        return jsonify({"success": False, "message": "Error interno."}), 500
